"""
El Cado - gestor de la pena
----------------------------
Backend Flask + SQLite.

Roles:
- El primer socio que se crea (cuando la base de datos esta vacia) se
  convierte automaticamente en administrador.
- Solo el administrador puede: anadir socios, dar de baja/reactivar
  socios, renombrar la pena, y anadir/borrar movimientos de caja
  (gastos e ingresos).
- Marcar una cuota como pagada solo lo pueden hacer el tesorero y el
  administrador (permiso manage_cuotas), previa comprobacion. El resto
  de socios ve las cuotas en modo lectura (quien ha pagado y quien no).
- Cualquier socio puede: reservar la pena, apuntarse a tareas, crear
  eventos de Tricount y participar en ellos, crear y gestionar listas de
  la compra compartidas (anadir o quitar productos de cualquiera), editar
  su propio perfil (incluido su nombre), y ver todo lo demas en modo
  lectura.

Para arrancar en local:
    pip install -r requirements.txt
    python app.py
"""
import os
import random
import sqlite3
import uuid
import json
from contextlib import closing
from datetime import date, datetime, timedelta
from io import BytesIO

from flask import Flask, g, jsonify, request, session, send_from_directory, render_template, send_file
from werkzeug.security import generate_password_hash, check_password_hash

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from PIL import Image, ImageOps
    PIL_OK = True
except ImportError:
    PIL_OK = False

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.abspath(os.environ.get("DATABASE_PATH", os.path.join(BASE_DIR, "pena.db")))
AVATARS_DIR = os.path.abspath(os.environ.get("AVATARS_DIR", os.path.join(BASE_DIR, "static", "avatars")))
TICKETS_DIR = os.path.abspath(os.environ.get("TICKETS_DIR", os.path.join(BASE_DIR, "static", "tickets")))

app = Flask(__name__)
# En produccion, define la variable de entorno SECRET_KEY con un valor propio.
app.secret_key = os.environ.get("SECRET_KEY", "cambia-esta-clave-en-produccion")

TAREAS_FIJAS = ["Compras", "Limpieza", "Tesoreria", "Carrozas", "Concursos", "Comidas", "Otros"]
ESTADOS_TICKET = ["pendiente", "en_curso", "hecho"]

PERMISSIONS = {
    "manage_roles": "Gestionar roles y permisos",
    "manage_socios": "Gestionar socios",
    "manage_config": "Gestionar configuracion de la pena",
    "view_finances": "Ver finanzas",
    "manage_finances": "Gestionar ingresos y gastos",
    "manage_cuotas": "Gestionar cuotas",
    "manage_bebidas": "Gestionar bebidas y precios",
    "manage_inventory": "Gestionar inventario",
    "manage_events": "Gestionar reuniones y reservas",
    "manage_tasks": "Asignar tareas",
    "export_data": "Exportar datos",
}
ALL_PERMISSIONS = list(PERMISSIONS)
DEFAULT_ROLES = [
    ("administrador", "Administrador", ALL_PERMISSIONS, 1),
    ("presidente", "Presidente", ["manage_config", "manage_socios", "view_finances", "manage_events", "manage_tasks", "export_data"], 1),
    ("vicepresidente", "Vicepresidente", ["view_finances", "manage_events", "manage_tasks", "export_data"], 1),
    ("tesorero", "Tesorero", ["view_finances", "manage_finances", "manage_cuotas", "manage_bebidas", "export_data"], 1),
    ("socio", "Socio", [], 1),
    ("otro", "Otro", [], 1),
]
SCHEMA = """
CREATE TABLE IF NOT EXISTS config (
    id INTEGER PRIMARY KEY CHECK (id = 1),
    nombre TEXT NOT NULL,
    cuota_mensual REAL NOT NULL
);
CREATE TABLE IF NOT EXISTS socios (
    id TEXT PRIMARY KEY,
    nombre TEXT NOT NULL,
    is_admin INTEGER NOT NULL DEFAULT 0,
    activo INTEGER NOT NULL DEFAULT 1,
    pin_hash TEXT,
    must_change_pin INTEGER NOT NULL DEFAULT 0
);
CREATE TABLE IF NOT EXISTS perfiles (
    socio_id TEXT PRIMARY KEY,
    telefono TEXT DEFAULT '',
    notas TEXT DEFAULT '',
    FOREIGN KEY (socio_id) REFERENCES socios(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS familiares (
    id TEXT PRIMARY KEY,
    socio_id TEXT NOT NULL,
    nombre TEXT NOT NULL,
    tipo TEXT NOT NULL,
    edad TEXT DEFAULT '',
    FOREIGN KEY (socio_id) REFERENCES socios(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS cuotas (
    id TEXT PRIMARY KEY,
    socio_id TEXT NOT NULL,
    year INTEGER NOT NULL,
    month INTEGER NOT NULL,
    importe REAL NOT NULL,
    pagado INTEGER NOT NULL DEFAULT 0,
    fecha TEXT,
    UNIQUE(socio_id, year, month)
);
CREATE TABLE IF NOT EXISTS reuniones (
    id TEXT PRIMARY KEY,
    fecha TEXT NOT NULL,
    hora_inicio TEXT,
    hora_fin TEXT,
    socio_id TEXT,
    evento TEXT NOT NULL,
    notas TEXT DEFAULT '',
    creado_en TEXT
);
CREATE TABLE IF NOT EXISTS asistencia (
    reunion_id TEXT NOT NULL,
    socio_id TEXT NOT NULL,
    PRIMARY KEY (reunion_id, socio_id)
);
CREATE TABLE IF NOT EXISTS inventario (
    id TEXT PRIMARY KEY,
    nombre TEXT NOT NULL,
    categoria TEXT NOT NULL,
    cantidad INTEGER NOT NULL DEFAULT 1,
    estado TEXT NOT NULL DEFAULT 'Bien',
    notas TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS movimientos (
    id TEXT PRIMARY KEY,
    tipo TEXT NOT NULL,
    categoria TEXT NOT NULL,
    concepto TEXT NOT NULL,
    importe REAL NOT NULL,
    fecha TEXT NOT NULL,
    socio_id TEXT,
    ticket INTEGER NOT NULL DEFAULT 0,
    ticket_ext TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS bebidas_precios (
    id TEXT PRIMARY KEY,
    nombre TEXT NOT NULL,
    unidad TEXT NOT NULL,
    precio_socio REAL NOT NULL,
    precio_no_socio REAL NOT NULL
);
CREATE TABLE IF NOT EXISTS bebidas_consumos (
    id TEXT PRIMARY KEY,
    fecha TEXT NOT NULL,
    consumidor TEXT NOT NULL,
    es_socio INTEGER NOT NULL,
    socio_id TEXT,
    bebida_id TEXT NOT NULL,
    cantidad INTEGER NOT NULL,
    importe REAL NOT NULL,
    pagado INTEGER NOT NULL DEFAULT 1
);
CREATE TABLE IF NOT EXISTS fiestas_gastos (
    id TEXT PRIMARY KEY,
    fecha TEXT NOT NULL,
    evento TEXT NOT NULL,
    concepto TEXT NOT NULL,
    importe REAL NOT NULL
);
CREATE TABLE IF NOT EXISTS reservas (
    id TEXT PRIMARY KEY,
    fecha TEXT NOT NULL,
    hora_inicio TEXT DEFAULT '',
    hora_fin TEXT DEFAULT '',
    socio_id TEXT NOT NULL,
    evento TEXT NOT NULL,
    notas TEXT DEFAULT '',
    creado_en TEXT
);
CREATE TABLE IF NOT EXISTS tareas_asignadas (
    id TEXT PRIMARY KEY,
    tarea TEXT NOT NULL,
    socio_id TEXT NOT NULL,
    UNIQUE(tarea, socio_id)
);
CREATE TABLE IF NOT EXISTS tareas_tickets (
    id TEXT PRIMARY KEY,
    tipo TEXT NOT NULL,
    responsable_id TEXT,
    estado TEXT NOT NULL DEFAULT 'pendiente',
    fecha TEXT,
    turno TEXT DEFAULT '',
    notas TEXT DEFAULT '',
    rotacion TEXT DEFAULT '',
    completado_en TEXT,
    creado_por TEXT,
    creado_en TEXT
);
CREATE TABLE IF NOT EXISTS tareas_ticket_socios (
    ticket_id TEXT NOT NULL,
    socio_id TEXT NOT NULL,
    PRIMARY KEY (ticket_id, socio_id),
    FOREIGN KEY (ticket_id) REFERENCES tareas_tickets(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS roles (
    id TEXT PRIMARY KEY,
    nombre TEXT NOT NULL UNIQUE,
    permisos TEXT NOT NULL DEFAULT '[]',
    es_sistema INTEGER NOT NULL DEFAULT 0
);
CREATE TABLE IF NOT EXISTS socio_roles (
    socio_id TEXT NOT NULL,
    rol_id TEXT NOT NULL,
    PRIMARY KEY (socio_id, rol_id),
    FOREIGN KEY (socio_id) REFERENCES socios(id) ON DELETE CASCADE,
    FOREIGN KEY (rol_id) REFERENCES roles(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS gastos_eventos (
    id TEXT PRIMARY KEY,
    nombre TEXT NOT NULL,
    fecha TEXT NOT NULL,
    notas TEXT DEFAULT '',
    creado_por TEXT,
    creado_en TEXT,
    oculto INTEGER NOT NULL DEFAULT 0
);
CREATE TABLE IF NOT EXISTS gastos_evento_participantes (
    evento_id TEXT NOT NULL,
    socio_id TEXT NOT NULL,
    PRIMARY KEY (evento_id, socio_id),
    FOREIGN KEY (evento_id) REFERENCES gastos_eventos(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS gastos_evento_pagos (
    id TEXT PRIMARY KEY,
    evento_id TEXT NOT NULL,
    pagador_id TEXT NOT NULL,
    concepto TEXT NOT NULL,
    importe REAL NOT NULL,
    fecha TEXT NOT NULL,
    beneficiarios TEXT NOT NULL DEFAULT '[]',
    FOREIGN KEY (evento_id) REFERENCES gastos_eventos(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS gastos_socios (
    id TEXT PRIMARY KEY,
    socio_id TEXT NOT NULL,
    tipo TEXT NOT NULL DEFAULT 'gasto',
    concepto TEXT NOT NULL,
    importe REAL NOT NULL,
    fecha TEXT NOT NULL,
    notas TEXT DEFAULT '',
    ticket INTEGER NOT NULL DEFAULT 0,
    ticket_ext TEXT DEFAULT '',
    abonado INTEGER NOT NULL DEFAULT 0,
    creado_en TEXT
);
CREATE TABLE IF NOT EXISTS listas_compra (
    id TEXT PRIMARY KEY,
    nombre TEXT NOT NULL,
    creado_por TEXT,
    creado_en TEXT
);
CREATE TABLE IF NOT EXISTS lista_compra_items (
    id TEXT PRIMARY KEY,
    lista_id TEXT NOT NULL,
    nombre TEXT NOT NULL,
    cantidad TEXT DEFAULT '',
    comprado INTEGER NOT NULL DEFAULT 0,
    anadido_por TEXT,
    comprado_por TEXT,
    creado_en TEXT,
    FOREIGN KEY (lista_id) REFERENCES listas_compra(id) ON DELETE CASCADE
);
"""


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    with closing(sqlite3.connect(DB_PATH)) as db:
        db.row_factory = sqlite3.Row
        db.executescript(SCHEMA)
        try:
            db.execute("ALTER TABLE socios ADD COLUMN pin_hash TEXT")
        except sqlite3.OperationalError:
            pass  # ya existia (base de datos creada con una version anterior)
        try:
            db.execute("ALTER TABLE socios ADD COLUMN must_change_pin INTEGER NOT NULL DEFAULT 0")
        except sqlite3.OperationalError:
            pass
        try:
            db.execute("ALTER TABLE movimientos ADD COLUMN socio_id TEXT")
        except sqlite3.OperationalError:
            pass
        try:
            db.execute("ALTER TABLE bebidas_consumos ADD COLUMN socio_id TEXT")
        except sqlite3.OperationalError:
            pass
        try:
            db.execute("ALTER TABLE bebidas_consumos ADD COLUMN pagado INTEGER NOT NULL DEFAULT 1")
        except sqlite3.OperationalError:
            pass
        try:
            db.execute("ALTER TABLE tareas_tickets ADD COLUMN completado_en TEXT")
        except sqlite3.OperationalError:
            pass
        try:
            db.execute("ALTER TABLE gastos_eventos ADD COLUMN oculto INTEGER NOT NULL DEFAULT 0")
        except sqlite3.OperationalError:
            pass
        try:
            db.execute("ALTER TABLE gastos_evento_pagos ADD COLUMN ticket INTEGER NOT NULL DEFAULT 0")
        except sqlite3.OperationalError:
            pass
        try:
            db.execute("ALTER TABLE gastos_socios ADD COLUMN tipo TEXT NOT NULL DEFAULT 'gasto'")
        except sqlite3.OperationalError:
            pass
        try:
            db.execute("ALTER TABLE movimientos ADD COLUMN ticket INTEGER NOT NULL DEFAULT 0")
        except sqlite3.OperationalError:
            pass
        try:
            db.execute("ALTER TABLE movimientos ADD COLUMN ticket_ext TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        try:
            db.execute("ALTER TABLE gastos_socios ADD COLUMN ticket_ext TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        # Migracion: actualizar tabla reuniones si existe con estructura antigua
        try:
            cur = db.execute("PRAGMA table_info(reuniones)")
            columns = [row[1] for row in cur.fetchall()]
            if "titulo" in columns and "evento" not in columns:
                # Migrar estructura antigua a nueva
                db.execute("ALTER TABLE reuniones RENAME TO reuniones_old")
                db.execute("""CREATE TABLE reuniones (
                    id TEXT PRIMARY KEY,
                    fecha TEXT NOT NULL,
                    hora_inicio TEXT,
                    hora_fin TEXT,
                    socio_id TEXT,
                    evento TEXT NOT NULL,
                    notas TEXT DEFAULT '',
                    creado_en TEXT
                )""")
                db.execute("""INSERT INTO reuniones (id, fecha, evento, notas, creado_en, socio_id)
                    SELECT id, fecha, titulo, notas, creado_en, creado_por FROM reuniones_old""")
                db.execute("DROP TABLE reuniones_old")
        except sqlite3.OperationalError:
            pass
        cur = db.execute("SELECT COUNT(*) AS n FROM config")
        if cur.fetchone()["n"] == 0:
            db.execute(
                "INSERT INTO config (id, nombre, cuota_mensual) VALUES (1, ?, ?)",
                ("El Cado", 45),
            )
        else:
            row = db.execute("SELECT cuota_mensual FROM config WHERE id = 1").fetchone()
            if row and (row["cuota_mensual"] in (None, "", 20, 20.0)):
                db.execute("UPDATE config SET cuota_mensual = ? WHERE id = 1", (45,))
        for role_id, nombre, permisos, es_sistema in DEFAULT_ROLES:
            db.execute(
                "INSERT OR IGNORE INTO roles (id, nombre, permisos, es_sistema) VALUES (?, ?, ?, ?)",
                (role_id, nombre, json.dumps(permisos), es_sistema),
            )
        # Las instalaciones previas solo tenian is_admin. Se traduce a un rol asignado.
        for socio in db.execute("SELECT id, is_admin FROM socios"):
            default_role = "administrador" if socio["is_admin"] else "socio"
            db.execute(
                "INSERT OR IGNORE INTO socio_roles (socio_id, rol_id) VALUES (?, ?)",
                (socio["id"], default_role),
            )
        # Migracion unica: las tareas del sistema antiguo (solo tipo + socio) pasan a ser tickets.
        if db.execute("SELECT COUNT(*) AS n FROM tareas_tickets").fetchone()["n"] == 0:
            for row in db.execute("SELECT tarea, socio_id FROM tareas_asignadas"):
                db.execute(
                    "INSERT INTO tareas_tickets (id, tipo, responsable_id, estado, fecha, turno, notas, rotacion, creado_por, creado_en) "
                    "VALUES (?, ?, ?, 'pendiente', NULL, '', '', '', NULL, ?)",
                    (uuid.uuid4().hex[:12], row["tarea"], row["socio_id"], datetime.now().isoformat(timespec="seconds")),
                )
        # Migracion unica: el responsable unico de cada ticket pasa a la tabla de socios multiples.
        if db.execute("SELECT COUNT(*) AS n FROM tareas_ticket_socios").fetchone()["n"] == 0:
            for row in db.execute("SELECT id, responsable_id FROM tareas_tickets WHERE responsable_id IS NOT NULL AND responsable_id != ''"):
                db.execute(
                    "INSERT OR IGNORE INTO tareas_ticket_socios (ticket_id, socio_id) VALUES (?, ?)",
                    (row["id"], row["responsable_id"]),
                )
        db.commit()


def new_id():
    return uuid.uuid4().hex[:12]


def now_iso():
    return datetime.now().isoformat(timespec="seconds")


def calcular_balances_evento(participantes, pagos):
    """balances[socio_id]: positivo = le deben, negativo = debe."""
    balances = {sid: 0.0 for sid in participantes}
    for pago in pagos:
        beneficiarios = [b for b in (pago["beneficiarios"] or participantes) if b in balances]
        if not beneficiarios:
            beneficiarios = list(participantes)
        share = pago["importe"] / len(beneficiarios)
        if pago["pagador_id"] in balances:
            balances[pago["pagador_id"]] += pago["importe"]
        for b in beneficiarios:
            balances[b] -= share
    return {sid: round(v, 2) for sid, v in balances.items()}


def calcular_transferencias_evento(balances):
    """Lista minima de pagos {de, a, importe} para saldar el evento."""
    deudores = sorted([[sid, -v] for sid, v in balances.items() if v < -0.005], key=lambda x: -x[1])
    acreedores = sorted([[sid, v] for sid, v in balances.items() if v > 0.005], key=lambda x: -x[1])
    transferencias = []
    i = j = 0
    while i < len(deudores) and j < len(acreedores):
        deudor_id, debe = deudores[i]
        acreedor_id, le_deben = acreedores[j]
        importe = min(debe, le_deben)
        if importe > 0.005:
            transferencias.append({"de": deudor_id, "a": acreedor_id, "importe": round(importe, 2)})
        deudores[i][1] -= importe
        acreedores[j][1] -= importe
        if deudores[i][1] <= 0.005:
            i += 1
        if acreedores[j][1] <= 0.005:
            j += 1
    return transferencias


DIAS_EN_MES = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]


def es_bisiesto(year):
    return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)


def siguiente_fecha_rotacion(fecha_iso, rotacion):
    if not fecha_iso:
        return None
    d = date.fromisoformat(fecha_iso)
    rot = (rotacion or "").strip().lower()
    if rot == "semanal":
        d = d + timedelta(days=7)
    elif rot == "quincenal":
        d = d + timedelta(days=14)
    elif rot == "mensual":
        month = d.month + 1
        year = d.year + (1 if month > 12 else 0)
        month = month if month <= 12 else 1
        dias_mes = DIAS_EN_MES[month - 1] + (1 if month == 2 and es_bisiesto(year) else 0)
        d = date(year, month, min(d.day, dias_mes))
    else:
        return None
    return d.isoformat()


def crear_siguiente_ocurrencia_tarea(db, ticket):
    nueva_fecha = siguiente_fecha_rotacion(ticket["fecha"], ticket["rotacion"])
    nuevo_id = new_id()
    socios = [r["socio_id"] for r in db.execute("SELECT socio_id FROM tareas_ticket_socios WHERE ticket_id = ?", (ticket["id"],))]
    db.execute(
        "INSERT INTO tareas_tickets (id, tipo, responsable_id, estado, fecha, notas, rotacion, completado_en, creado_por, creado_en) "
        "VALUES (?,?,?,'pendiente',?,?,?,NULL,?,?)",
        (
            nuevo_id, ticket["tipo"], socios[0] if socios else None, nueva_fecha,
            ticket["notas"], ticket["rotacion"], ticket["creado_por"], now_iso(),
        ),
    )
    if socios:
        db.executemany("INSERT OR IGNORE INTO tareas_ticket_socios (ticket_id, socio_id) VALUES (?, ?)", [(nuevo_id, s) for s in socios])
    return nuevo_id


def current_socio_id():
    return session.get("socio_id")


def require_login():
    """Devuelve el id del socio en sesion, o None. Limpia la sesion si el socio ya no existe."""
    sid = current_socio_id()
    if not sid:
        return None
    db = get_db()
    row = db.execute("SELECT id FROM socios WHERE id = ?", (sid,)).fetchone()
    if not row:
        session.pop("socio_id", None)
        return None
    return sid


def is_admin_user(sid):
    return has_permission(sid, "manage_roles")


def permissions_for_user(sid):
    if not sid:
        return set()
    db = get_db()
    rows = db.execute(
        "SELECT r.permisos FROM roles r JOIN socio_roles sr ON sr.rol_id = r.id WHERE sr.socio_id = ?",
        (sid,),
    ).fetchall()
    permisos = set()
    for row in rows:
        try:
            permisos.update(json.loads(row["permisos"]))
        except (TypeError, ValueError, json.JSONDecodeError):
            continue
    return permisos


def has_permission(sid, permission):
    return permission in permissions_for_user(sid)


def require_permission(permission):
    sid = require_login()
    return sid if has_permission(sid, permission) else None


def require_admin():
    sid = require_login()
    return sid if is_admin_user(sid) else None


def err(msg, code=403):
    return jsonify({"error": msg}), code


# ---------------------------------------------------------------- paginas --
def asset_version():
    """Cambia cada vez que se edita app.js o style.css, para que el movil
    (sobre todo la app instalada como PWA) no se quede con una copia vieja
    en cache y siga viendo comportamiento antiguo tras una actualizacion."""
    try:
        mtimes = [
            os.path.getmtime(os.path.join(BASE_DIR, "static", "app.js")),
            os.path.getmtime(os.path.join(BASE_DIR, "static", "style.css")),
        ]
        return str(int(max(mtimes)))
    except OSError:
        return "1"


@app.route("/")
def index():
    return render_template("index.html", asset_version=asset_version())


@app.route("/static/<path:path>")
def static_files(path):
    response = send_from_directory(os.path.join(BASE_DIR, "static"), path)
    if path.startswith("avatars/") or path.startswith("tickets/"):
        # Estos archivos se sobrescriben con el mismo nombre al cambiar la foto
        # o el ticket. Algunos navegadores moviles (sobre todo Safari/iOS en
        # apps instaladas) reutilizan la imagen en cache aunque cambie el
        # parametro ?v= de la URL, asi que se fuerza a no guardarlas nunca.
        response.headers["Cache-Control"] = "no-store, must-revalidate"
    return response


def valid_pin(pin):
    return bool(pin) and len(pin) == 4 and pin.isdigit()


def random_pin():
    return f"{random.randint(0, 9999):04d}"


# ---------------------------------------------------------------- sesion --
@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json(force=True)
    socio_id = data.get("socio_id")
    pin = (data.get("pin") or "").strip()
    db = get_db()
    row = db.execute("SELECT id, activo, pin_hash FROM socios WHERE id = ?", (socio_id,)).fetchone()
    if not row:
        return err("Socio no encontrado", 404)
    if not row["activo"]:
        return err("Este socio esta dado de baja", 403)
    if row["pin_hash"]:
        if not pin or not check_password_hash(row["pin_hash"], pin):
            return err("PIN incorrecto.", 403)
    # Si el socio todavia no tiene PIN configurado (cuentas antiguas), se deja
    # entrar sin comprobarlo; se le anima a crear uno desde "Mi perfil".
    session["socio_id"] = socio_id
    return jsonify({"ok": True})


@app.route("/api/logout", methods=["POST"])
def logout():
    session.pop("socio_id", None)
    return jsonify({"ok": True})


# ------------------------------------------------------------- estado full --
@app.route("/api/state")
def state():
    db = get_db()
    sid = require_login()

    config = dict(db.execute("SELECT * FROM config WHERE id = 1").fetchone())

    socios_rows = db.execute("SELECT * FROM socios ORDER BY nombre").fetchall()
    roles = []
    for r in db.execute("SELECT * FROM roles ORDER BY nombre"):
        role = dict(r)
        try:
            role["permisos"] = json.loads(role["permisos"])
        except (TypeError, ValueError, json.JSONDecodeError):
            role["permisos"] = []
        roles.append(role)
    roles_por_socio = {}
    for r in db.execute("SELECT socio_id, rol_id FROM socio_roles"):
        roles_por_socio.setdefault(r["socio_id"], []).append(r["rol_id"])

    socios = []
    for r in socios_rows:
        d = dict(r)
        d["tiene_pin"] = bool(d.pop("pin_hash", None))
        d["must_change_pin"] = bool(d.get("must_change_pin", 0))
        d["roles"] = roles_por_socio.get(d["id"], ["socio"])
        socios.append(d)

    perfiles = {}
    for r in db.execute("SELECT * FROM perfiles"):
        perfiles[r["socio_id"]] = {"telefono": r["telefono"], "notas": r["notas"], "familia": []}
    for r in db.execute("SELECT * FROM familiares"):
        perfiles.setdefault(r["socio_id"], {"telefono": "", "notas": "", "familia": []})
        perfiles[r["socio_id"]]["familia"].append(
            {"id": r["id"], "nombre": r["nombre"], "tipo": r["tipo"], "edad": r["edad"]}
        )

    can_view_finances = has_permission(sid, "view_finances")
    can_manage_cuotas = has_permission(sid, "manage_cuotas")
    if can_view_finances or can_manage_cuotas:
        cuotas = [dict(r) for r in db.execute("SELECT * FROM cuotas")]
    elif sid:
        cuotas = [dict(r) for r in db.execute("SELECT * FROM cuotas WHERE socio_id = ?", (sid,))]
    else:
        cuotas = []

    reuniones = []
    for r in db.execute("SELECT * FROM reuniones ORDER BY fecha DESC"):
        asistentes = [
            a["socio_id"]
            for a in db.execute("SELECT socio_id FROM asistencia WHERE reunion_id = ?", (r["id"],))
        ]
        row = dict(r)
        row["asistentes"] = asistentes
        reuniones.append(row)

    inventario = [dict(r) for r in db.execute("SELECT * FROM inventario ORDER BY categoria, nombre")]
    consumo_cols = (
        "SELECT bc.*, bp.nombre AS bebida_nombre FROM bebidas_consumos bc "
        "LEFT JOIN bebidas_precios bp ON bp.id = bc.bebida_id "
    )
    bebidas_precios = [dict(r) for r in db.execute("SELECT * FROM bebidas_precios ORDER BY nombre")]
    fiestas_gastos = [dict(r) for r in db.execute("SELECT * FROM fiestas_gastos ORDER BY fecha DESC")]
    if can_view_finances or has_permission(sid, "manage_finances"):
        movimientos = [dict(r) for r in db.execute("SELECT m.*, s.nombre AS socio_nombre FROM movimientos m LEFT JOIN socios s ON s.id = m.socio_id ORDER BY fecha DESC")]
        bebidas_consumos = [dict(r) for r in db.execute(consumo_cols + "ORDER BY bc.fecha DESC LIMIT 80")]
    else:
        movimientos = []
        bebidas_consumos = (
            [dict(r) for r in db.execute(consumo_cols + "WHERE bc.socio_id = ? ORDER BY bc.fecha DESC", (sid,))]
            if sid else []
        )
    reservas = [dict(r) for r in db.execute("SELECT * FROM reservas ORDER BY fecha")]

    listas_compra = []
    for r in db.execute("SELECT * FROM listas_compra ORDER BY creado_en DESC"):
        lista = dict(r)
        lista["items"] = [
            dict(i) for i in db.execute(
                "SELECT * FROM lista_compra_items WHERE lista_id = ? ORDER BY comprado, creado_en",
                (lista["id"],),
            )
        ]
        listas_compra.append(lista)

    gastos_eventos = []
    for r in db.execute("SELECT * FROM gastos_eventos ORDER BY fecha DESC"):
        evento = dict(r)
        participantes = [
            p["socio_id"]
            for p in db.execute(
                "SELECT socio_id FROM gastos_evento_participantes WHERE evento_id = ?", (evento["id"],)
            )
        ]
        pagos = []
        for p in db.execute(
            "SELECT * FROM gastos_evento_pagos WHERE evento_id = ? ORDER BY fecha DESC", (evento["id"],)
        ):
            pago = dict(p)
            try:
                pago["beneficiarios"] = json.loads(pago["beneficiarios"])
            except (TypeError, ValueError, json.JSONDecodeError):
                pago["beneficiarios"] = []
            pagos.append(pago)
        balances = calcular_balances_evento(participantes, pagos)
        evento["participantes"] = participantes
        evento["pagos"] = pagos
        evento["balances"] = balances
        evento["transferencias"] = calcular_transferencias_evento(balances)
        evento["total"] = round(sum(p["importe"] for p in pagos), 2)
        gastos_eventos.append(evento)

    tareas_tickets = [dict(r) for r in db.execute("SELECT * FROM tareas_tickets ORDER BY (fecha IS NULL), fecha, tipo")]
    tareas_socios_de = {}
    for r in db.execute("SELECT * FROM tareas_ticket_socios"):
        tareas_socios_de.setdefault(r["ticket_id"], []).append(r["socio_id"])
    for t in tareas_tickets:
        t["socios_ids"] = tareas_socios_de.get(t["id"], [])

    gastos_socios = [
        dict(r) for r in db.execute(
            "SELECT gs.*, s.nombre AS socio_nombre FROM gastos_socios gs "
            "LEFT JOIN socios s ON s.id = gs.socio_id ORDER BY gs.fecha DESC"
        )
    ]

    return jsonify(
        {
            "config": config,
            "roles": roles,
            "permission_labels": PERMISSIONS,
            "permissions": sorted(permissions_for_user(sid)),
            "socios": socios,
            "perfiles": perfiles,
            "cuotas": cuotas,
            "reuniones": reuniones,
            "inventario": inventario,
            "movimientos": movimientos,
            "bebidas_precios": bebidas_precios,
            "bebidas_consumos": bebidas_consumos,
            "fiestas_gastos": fiestas_gastos,
            "reservas": reservas,
            "listas_compra": listas_compra,
            "gastos_eventos": gastos_eventos,
            "gastos_socios": gastos_socios,
            "tareas_tickets": tareas_tickets,
            "tareas_fijas": TAREAS_FIJAS,
            "estados_ticket": ESTADOS_TICKET,
            "current_user": sid,
            "is_admin": is_admin_user(sid),
        }
    )


# -------------------------------------------------------------- config (solo admin) --
@app.route("/api/config", methods=["POST"])
def update_config():
    if not require_permission("manage_config"):
        return err("Solo el administrador puede renombrar la pena o cambiar la cuota.")
    data = request.get_json(force=True)
    db = get_db()
    try:
        cuota = float(data.get("cuota_mensual", 45))
    except (TypeError, ValueError):
        cuota = 45
    db.execute(
        "UPDATE config SET nombre = ?, cuota_mensual = ? WHERE id = 1",
        ((data.get("nombre") or "El Cado").strip(), cuota),
    )
    db.commit()
    return jsonify({"ok": True})


# -------------------------------------------------------------- socios --
@app.route("/api/socios", methods=["POST"])
def add_socio():
    data = request.get_json(force=True)
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return err("Nombre requerido", 400)
    pin = (data.get("pin") or "").strip()

    db = get_db()
    total = db.execute("SELECT COUNT(*) AS n FROM socios").fetchone()["n"]

    if total == 0:
        # Bootstrap: el primer socio que se crea es el administrador y DEBE fijar un PIN.
        if not valid_pin(pin):
            return err("El PIN debe tener 4 digitos.", 400)
        is_admin = 1
        pin_generado = None
        must_change = 0  # el propio admin ha elegido su PIN, no hace falta forzar cambio
    else:
        if not require_permission("manage_socios"):
            return err("Solo el administrador puede anadir socios nuevos.")
        is_admin = 0
        must_change = 1  # el PIN lo ha puesto el administrador: se le obligara a cambiarlo
        if pin:
            if not valid_pin(pin):
                return err("El PIN debe tener 4 digitos.", 400)
            pin_generado = None
        else:
            pin = random_pin()
            pin_generado = pin

    sid = new_id()
    db.execute(
        "INSERT INTO socios (id, nombre, is_admin, activo, pin_hash, must_change_pin) VALUES (?, ?, ?, 1, ?, ?)",
        (sid, nombre, is_admin, generate_password_hash(pin), must_change),
    )
    db.execute("INSERT INTO perfiles (socio_id, telefono, notas) VALUES (?, '', '')", (sid,))
    db.execute("INSERT INTO socio_roles (socio_id, rol_id) VALUES (?, ?)", (sid, 'administrador' if is_admin else 'socio'))
    db.commit()

    if total == 0:
        session["socio_id"] = sid  # login automatico del primer administrador

    return jsonify({"ok": True, "id": sid, "is_admin": bool(is_admin), "pin_generado": pin_generado})


@app.route("/api/socios/<sid>/reset-pin", methods=["POST"])
def reset_pin(sid):
    if not require_permission("manage_socios"):
        return err("Solo el administrador puede restablecer un PIN.")
    db = get_db()
    if not db.execute("SELECT 1 FROM socios WHERE id = ?", (sid,)).fetchone():
        return err("Socio no encontrado", 404)
    nuevo_pin = random_pin()
    db.execute(
        "UPDATE socios SET pin_hash = ?, must_change_pin = 1 WHERE id = ?",
        (generate_password_hash(nuevo_pin), sid),
    )
    db.commit()
    return jsonify({"ok": True, "pin": nuevo_pin})


@app.route("/api/socios/<sid>/activo", methods=["POST"])
def toggle_activo(sid):
    if not require_permission("manage_socios"):
        return err("Solo el administrador puede dar de alta o de baja a un socio.")
    db = get_db()
    row = db.execute("SELECT activo FROM socios WHERE id = ?", (sid,)).fetchone()
    if not row:
        return err("Socio no encontrado", 404)
    nuevo = 0 if row["activo"] else 1
    db.execute("UPDATE socios SET activo = ? WHERE id = ?", (nuevo, sid))
    db.commit()
    return jsonify({"ok": True, "activo": bool(nuevo)})


@app.route("/api/socios/<sid>", methods=["DELETE"])
def delete_socio(sid):
    if not require_permission("manage_socios"):
        return err("Solo el administrador puede eliminar un socio.")
    db = get_db()
    row = db.execute("SELECT id, is_admin FROM socios WHERE id = ?", (sid,)).fetchone()
    if not row:
        return err("Socio no encontrado", 404)
    if row["is_admin"]:
        return err("No se puede eliminar a un administrador.", 403)

    db.execute("DELETE FROM cuotas WHERE socio_id = ?", (sid,))
    db.execute("DELETE FROM reservas WHERE socio_id = ?", (sid,))
    db.execute("DELETE FROM asistencia WHERE socio_id = ?", (sid,))
    db.execute("DELETE FROM tareas_asignadas WHERE socio_id = ?", (sid,))
    db.execute("DELETE FROM socios WHERE id = ?", (sid,))
    db.commit()

    avatar_path = os.path.join(AVATARS_DIR, f"{sid}.jpg")
    try:
        if os.path.exists(avatar_path):
            os.remove(avatar_path)
    except OSError:
        pass

    return jsonify({"ok": True})


# -------------------------------------------------------------- roles y permisos --
@app.route("/api/socios/<sid>/roles", methods=["POST"])
def update_socio_roles(sid):
    if not require_permission("manage_roles"):
        return err("No tienes permiso para gestionar roles.")
    data = request.get_json(force=True)
    role_ids = data.get("roles") or []
    if not isinstance(role_ids, list) or not role_ids:
        return err("Cada socio debe tener al menos un rol.", 400)
    db = get_db()
    if not db.execute("SELECT 1 FROM socios WHERE id = ?", (sid,)).fetchone():
        return err("Socio no encontrado", 404)
    valid_ids = {r["id"] for r in db.execute("SELECT id FROM roles")}
    if any(role_id not in valid_ids for role_id in role_ids):
        return err("Uno de los roles no existe.", 400)
    # Evita que alguien se quite a si mismo eel ultimo rol con permiso para administrar roles.
    if sid == current_socio_id():
        permitted = set()
        for role in db.execute("SELECT permisos FROM roles WHERE id IN ({})".format(",".join("?" for _ in role_ids)), role_ids):
            permitted.update(json.loads(role["permisos"]))
        if "manage_roles" not in permitted:
            return err("No puedes quitarte a ti mismo el permiso de gestionar roles.", 400)
    db.execute("DELETE FROM socio_roles WHERE socio_id = ?", (sid,))
    db.executemany("INSERT INTO socio_roles (socio_id, rol_id) VALUES (?, ?)", [(sid, role_id) for role_id in role_ids])
    db.execute("UPDATE socios SET is_admin = ? WHERE id = ?", (int("administrador" in role_ids), sid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/roles", methods=["POST"])
def create_role():
    if not require_permission("manage_roles"):
        return err("No tienes permiso para gestionar roles.")
    data = request.get_json(force=True)
    nombre = (data.get("nombre") or "").strip()
    permisos = data.get("permisos") or []
    if not nombre:
        return err("El nombre del rol es obligatorio.", 400)
    if not isinstance(permisos, list) or any(p not in PERMISSIONS for p in permisos):
        return err("La lista de permisos no es valida.", 400)
    db = get_db()
    role_id = new_id()
    try:
        db.execute("INSERT INTO roles (id, nombre, permisos) VALUES (?, ?, ?)", (role_id, nombre, json.dumps(permisos)))
        db.commit()
    except sqlite3.IntegrityError:
        return err("Ya existe un rol con ese nombre.", 400)
    return jsonify({"ok": True, "id": role_id})


@app.route("/api/roles/<role_id>", methods=["POST"])
def update_role(role_id):
    if not require_permission("manage_roles"):
        return err("No tienes permiso para gestionar roles.")
    data = request.get_json(force=True)
    permisos = data.get("permisos") or []
    if not isinstance(permisos, list) or any(p not in PERMISSIONS for p in permisos):
        return err("La lista de permisos no es valida.", 400)
    if role_id == "administrador" and "manage_roles" not in permisos:
        return err("El rol Administrador debe conservar la gestion de roles.", 400)
    db = get_db()
    if not db.execute("SELECT 1 FROM roles WHERE id = ?", (role_id,)).fetchone():
        return err("Rol no encontrado", 404)
    db.execute("UPDATE roles SET permisos = ? WHERE id = ?", (json.dumps(permisos), role_id))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/roles/<role_id>", methods=["DELETE"])
def delete_role(role_id):
    if not require_permission("manage_roles"):
        return err("No tienes permiso para gestionar roles.")
    db = get_db()
    row = db.execute("SELECT es_sistema FROM roles WHERE id = ?", (role_id,)).fetchone()
    if not row:
        return err("Rol no encontrado", 404)
    if row["es_sistema"]:
        return err("Los roles del sistema no se pueden borrar.", 400)
    huerfanos = db.execute(
        "SELECT sr.socio_id FROM socio_roles sr WHERE sr.rol_id = ? "
        "AND (SELECT COUNT(*) FROM socio_roles sr2 WHERE sr2.socio_id = sr.socio_id) = 1",
        (role_id,),
    ).fetchall()
    if huerfanos:
        return err("Hay socios que solo tienen este rol: asignales otro antes de borrarlo.", 400)
    db.execute("DELETE FROM roles WHERE id = ?", (role_id,))
    db.commit()
    return jsonify({"ok": True})

# -------------------------------------------------------------- perfil (solo el propio) --
@app.route("/api/perfil", methods=["POST"])
def update_perfil():
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    db = get_db()

    nombre = (data.get("nombre") or "").strip()
    if nombre:
        db.execute("UPDATE socios SET nombre = ? WHERE id = ?", (nombre, sid))

    pin = (data.get("pin") or "").strip()
    if pin:
        if not valid_pin(pin):
            return err("El PIN debe tener 4 digitos.", 400)
        db.execute("UPDATE socios SET pin_hash = ?, must_change_pin = 0 WHERE id = ?", (generate_password_hash(pin), sid))

    db.execute(
        "UPDATE perfiles SET telefono = ?, notas = ? WHERE socio_id = ?",
        (data.get("telefono", ""), data.get("notas", ""), sid),
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/perfil/pin", methods=["POST"])
def change_own_pin():
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    pin = (data.get("pin") or "").strip()
    if not valid_pin(pin):
        return err("El PIN debe tener 4 digitos.", 400)
    db = get_db()
    db.execute("UPDATE socios SET pin_hash = ?, must_change_pin = 0 WHERE id = ?", (generate_password_hash(pin), sid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/familiares", methods=["POST"])
def add_familiar():
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return err("Nombre requerido", 400)
    db = get_db()
    fid = new_id()
    db.execute(
        "INSERT INTO familiares (id, socio_id, nombre, tipo, edad) VALUES (?, ?, ?, ?, ?)",
        (fid, sid, nombre, data.get("tipo", "Otro"), str(data.get("edad", ""))),
    )
    db.commit()
    return jsonify({"ok": True, "id": fid})


@app.route("/api/familiares/<fid>", methods=["DELETE"])
def delete_familiar(fid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    db.execute("DELETE FROM familiares WHERE id = ? AND socio_id = ?", (fid, sid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/socios/<sid>/foto", methods=["POST"])
def upload_foto(sid):
    sid_session = require_login()
    if not sid_session:
        return err("No has iniciado sesion.", 401)
    if sid_session != sid and not has_permission(sid_session, "manage_socios"):
        return err("No puedes cambiar la foto de otro socio.")
    if not PIL_OK:
        return err(
            "Falta instalar una dependencia en el servidor: ejecuta "
            "'pip install -r requirements.txt' (o 'pip install Pillow') "
            "y reinicia la aplicacion.",
            500,
        )
    if "foto" not in request.files or request.files["foto"].filename == "":
        return err("No se ha recibido ninguna imagen.", 400)

    db = get_db()
    if not db.execute("SELECT 1 FROM socios WHERE id = ?", (sid,)).fetchone():
        return err("Socio no encontrado", 404)

    try:
        file = request.files["foto"]
        raw = file.read()
        try:
            im = Image.open(BytesIO(raw))
            im.load()
        except Exception:
            return err(
                "No se pudo leer esa imagen. Prueba con un archivo .jpg o .png "
                "(si es una foto de iPhone en formato HEIC, conviertela a JPG antes de subirla).",
                400,
            )
        im = ImageOps.exif_transpose(im).convert("RGB")
        w, h = im.size
        side = min(w, h)
        left, top = (w - side) // 2, (h - side) // 2
        im = im.crop((left, top, left + side, top + side)).resize((320, 320), Image.LANCZOS)
        os.makedirs(AVATARS_DIR, exist_ok=True)
        im.save(os.path.join(AVATARS_DIR, f"{sid}.jpg"), "JPEG", quality=85)
    except Exception as e:
        return err(f"No se pudo procesar la imagen: {e}", 400)

    return jsonify({"ok": True})


# -------------------------------------------------------------- cuotas --
@app.route("/api/cuota/toggle", methods=["POST"])
def toggle_cuota():
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    socio_id, year, month = data.get("socio_id"), int(data.get("year")), int(data.get("month"))

    if not has_permission(sid, "manage_cuotas"):
        return err("Solo el tesorero o el administrador pueden marcar una cuota como pagada, previa comprobacion.")

    db = get_db()
    row = db.execute(
        "SELECT * FROM cuotas WHERE socio_id = ? AND year = ? AND month = ?",
        (socio_id, year, month),
    ).fetchone()
    config = db.execute("SELECT cuota_mensual FROM config WHERE id = 1").fetchone()
    if row:
        nuevo_estado = 0 if row["pagado"] else 1
        db.execute(
            "UPDATE cuotas SET pagado = ?, fecha = ? WHERE id = ?",
            (nuevo_estado, date.today().isoformat() if nuevo_estado else None, row["id"]),
        )
    else:
        db.execute(
            "INSERT INTO cuotas (id, socio_id, year, month, importe, pagado, fecha) VALUES (?,?,?,?,?,1,?)",
            (new_id(), socio_id, year, month, config["cuota_mensual"], date.today().isoformat()),
        )
    db.commit()
    return jsonify({"ok": True})


# -------------------------------------------------------------- reuniones --
@app.route("/api/reuniones", methods=["POST"])
def add_reunion():
    sid = require_permission("manage_events")
    data = request.get_json(force=True)
    db = get_db()
    rid = new_id()
    db.execute(
        "INSERT INTO reuniones (id, fecha, hora_inicio, hora_fin, socio_id, evento, notas, creado_en) VALUES (?,?,?,?,?,?,?,?)",
        (
            rid,
            data.get("fecha"),
            (data.get("hora_inicio") or "").strip(),
            (data.get("hora_fin") or "").strip(),
            sid,
            (data.get("evento") or "").strip(),
            (data.get("notas") or "").strip(),
            now_iso(),
        ),
    )
    db.commit()
    return jsonify({"ok": True, "id": rid})


@app.route("/api/reuniones/<rid>", methods=["DELETE"])
def delete_reunion(rid):
    if not require_permission("manage_events"):
        return err("No tienes permiso para gestionar reuniones.")
    db = get_db()
    db.execute("DELETE FROM reuniones WHERE id = ?", (rid,))
    db.execute("DELETE FROM asistencia WHERE reunion_id = ?", (rid,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/reuniones/<rid>/asistencia", methods=["POST"])
def toggle_asistencia(rid):
    data = request.get_json(force=True)
    socio_id = data.get("socio_id")
    db = get_db()
    row = db.execute("SELECT 1 FROM asistencia WHERE reunion_id = ? AND socio_id = ?", (rid, socio_id)).fetchone()
    if row:
        db.execute("DELETE FROM asistencia WHERE reunion_id = ? AND socio_id = ?", (rid, socio_id))
    else:
        db.execute("INSERT INTO asistencia (reunion_id, socio_id) VALUES (?, ?)", (rid, socio_id))
    db.commit()
    return jsonify({"ok": True})


# -------------------------------------------------------------- inventario --
@app.route("/api/inventario", methods=["POST"])
def add_inventario():
    if not require_permission("manage_inventory"):
        return err("No tienes permiso para gestionar inventario.")
    data = request.get_json(force=True)
    db = get_db()
    iid = new_id()
    db.execute(
        "INSERT INTO inventario (id, nombre, categoria, cantidad, estado, notas) VALUES (?,?,?,?,?,?)",
        (
            iid,
            (data.get("nombre") or "").strip(),
            data.get("categoria", "Otros"),
            int(data.get("cantidad") or 1),
            data.get("estado", "Bien"),
            (data.get("notas") or "").strip(),
        ),
    )
    db.commit()
    return jsonify({"ok": True, "id": iid})


@app.route("/api/inventario/<iid>", methods=["POST"])
def update_inventario(iid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    if not db.execute("SELECT 1 FROM inventario WHERE id = ?", (iid,)).fetchone():
        return err("Material no encontrado", 404)
    data = request.get_json(force=True)
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return err("El nombre es obligatorio.", 400)
    db.execute(
        "UPDATE inventario SET nombre = ?, categoria = ?, cantidad = ?, estado = ?, notas = ? WHERE id = ?",
        (
            nombre,
            data.get("categoria", "Otros"),
            int(data.get("cantidad") or 1),
            data.get("estado", "Bien"),
            (data.get("notas") or "").strip(),
            iid,
        ),
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/inventario/<iid>", methods=["DELETE"])
def delete_inventario(iid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    db.execute("DELETE FROM inventario WHERE id = ?", (iid,))
    db.commit()
    return jsonify({"ok": True})


# -------------------------------------------------------------- movimientos (solo admin) --
@app.route("/api/movimientos", methods=["POST"])
def add_movimiento():
    if not require_permission("manage_finances"):
        return err("Solo el administrador puede anadir gastos o ingresos.")
    data = request.get_json(force=True)
    db = get_db()
    socio_id = data.get("socio_id") or None
    if socio_id:
        if not db.execute("SELECT 1 FROM socios WHERE id = ?", (socio_id,)).fetchone():
            return err("Socio no encontrado", 404)
    mid = new_id()
    db.execute(
        "INSERT INTO movimientos (id, tipo, categoria, concepto, importe, fecha, socio_id) VALUES (?,?,?,?,?,?,?)",
        (
            mid,
            data.get("tipo", "gasto"),
            data.get("categoria", "Otros"),
            (data.get("concepto") or "").strip(),
            float(data.get("importe") or 0),
            data.get("fecha"),
            socio_id,
        ),
    )
    db.commit()
    return jsonify({"ok": True, "id": mid})


@app.route("/api/movimientos/<mid>", methods=["POST"])
def update_movimiento(mid):
    if not require_permission("manage_finances"):
        return err("Solo el administrador o el tesorero pueden modificar gastos o ingresos.")
    db = get_db()
    if not db.execute("SELECT 1 FROM movimientos WHERE id = ?", (mid,)).fetchone():
        return err("Movimiento no encontrado", 404)
    data = request.get_json(force=True)
    socio_id = data.get("socio_id") or None
    if socio_id:
        if not db.execute("SELECT 1 FROM socios WHERE id = ?", (socio_id,)).fetchone():
            return err("Socio no encontrado", 404)
    db.execute(
        "UPDATE movimientos SET tipo=?, categoria=?, concepto=?, importe=?, fecha=?, socio_id=? WHERE id=?",
        (
            data.get("tipo", "gasto"),
            data.get("categoria", "Otros"),
            (data.get("concepto") or "").strip(),
            float(data.get("importe") or 0),
            data.get("fecha"),
            socio_id,
            mid,
        ),
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/movimientos/<mid>", methods=["DELETE"])
def delete_movimiento(mid):
    if not require_permission("manage_finances"):
        return err("Solo el administrador puede borrar gastos o ingresos.")
    db = get_db()
    db.execute("DELETE FROM movimientos WHERE id = ?", (mid,))
    db.commit()
    for ext in ("jpg", "pdf"):
        ticket_path = os.path.join(TICKETS_DIR, f"mov-{mid}.{ext}")
        try:
            if os.path.exists(ticket_path):
                os.remove(ticket_path)
        except OSError:
            pass
    return jsonify({"ok": True})


@app.route("/api/movimientos/<mid>/ticket", methods=["POST"])
def upload_ticket_movimiento(mid):
    if not require_permission("manage_finances"):
        return err("Solo el administrador o el tesorero pueden subir el ticket o factura.")
    db = get_db()
    if not db.execute("SELECT 1 FROM movimientos WHERE id = ?", (mid,)).fetchone():
        return err("Movimiento no encontrado", 404)
    if "ticket" not in request.files or request.files["ticket"].filename == "":
        return err("No se ha recibido ningun archivo.", 400)
    ext, error = _guardar_ticket_o_factura(request.files["ticket"], TICKETS_DIR, f"mov-{mid}")
    if error:
        return error
    db.execute("UPDATE movimientos SET ticket = 1, ticket_ext = ? WHERE id = ?", (ext, mid))
    db.commit()
    return jsonify({"ok": True, "ext": ext})


@app.route("/api/movimientos/<mid>/ticket", methods=["DELETE"])
def delete_ticket_movimiento(mid):
    if not require_permission("manage_finances"):
        return err("Solo el administrador o el tesorero pueden quitar el ticket o factura.")
    db = get_db()
    if not db.execute("SELECT 1 FROM movimientos WHERE id = ?", (mid,)).fetchone():
        return err("Movimiento no encontrado", 404)
    for ext in ("jpg", "pdf"):
        ticket_path = os.path.join(TICKETS_DIR, f"mov-{mid}.{ext}")
        try:
            if os.path.exists(ticket_path):
                os.remove(ticket_path)
        except OSError:
            pass
    db.execute("UPDATE movimientos SET ticket = 0, ticket_ext = '' WHERE id = ?", (mid,))
    db.commit()
    return jsonify({"ok": True})


# -------------------------------------------------------------- bebidas: precios --
@app.route("/api/bebidas/precios", methods=["POST"])
def add_bebida_precio():
    if not require_permission("manage_bebidas"):
        return err("No tienes permiso para gestionar bebidas.")
    data = request.get_json(force=True)
    db = get_db()
    bid = new_id()
    db.execute(
        "INSERT INTO bebidas_precios (id, nombre, unidad, precio_socio, precio_no_socio) VALUES (?,?,?,?,?)",
        (
            bid,
            (data.get("nombre") or "").strip(),
            (data.get("unidad") or "").strip(),
            float(data.get("precio_socio") or 0),
            float(data.get("precio_no_socio") or 0),
        ),
    )
    db.commit()
    return jsonify({"ok": True, "id": bid})


@app.route("/api/bebidas/precios/<bid>", methods=["DELETE"])
def delete_bebida_precio(bid):
    if not require_permission("manage_bebidas"):
        return err("No tienes permiso para gestionar bebidas.")
    db = get_db()
    db.execute("DELETE FROM bebidas_precios WHERE id = ?", (bid,))
    db.commit()
    return jsonify({"ok": True})


# -------------------------------------------------------------- bebidas: consumos --
@app.route("/api/bebidas/consumos", methods=["POST"])
def add_consumo():
    if not require_login():
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    db = get_db()
    bebida = db.execute("SELECT * FROM bebidas_precios WHERE id = ?", (data.get("bebida_id"),)).fetchone()
    if not bebida:
        return err("Bebida no encontrada", 404)

    es_socio = bool(data.get("es_socio"))
    cantidad = int(data.get("cantidad") or 1)
    socio_id = None
    if es_socio:
        socio_id = data.get("socio_id") or None
        consumidor_row = db.execute("SELECT nombre FROM socios WHERE id = ?", (socio_id,)).fetchone()
        consumidor = consumidor_row["nombre"] if consumidor_row else "Socio"
        precio_unit = bebida["precio_socio"]
    else:
        consumidor = (data.get("nombre_invitado") or "Invitado").strip()
        precio_unit = bebida["precio_no_socio"]
    pagado = 1 if data.get("pagado", True) else 0

    cid = new_id()
    db.execute(
        "INSERT INTO bebidas_consumos (id, fecha, consumidor, es_socio, socio_id, bebida_id, cantidad, importe, pagado) VALUES (?,?,?,?,?,?,?,?,?)",
        (cid, date.today().isoformat(), consumidor, int(es_socio), socio_id, bebida["id"], cantidad, precio_unit * cantidad, pagado),
    )
    db.commit()
    return jsonify({"ok": True, "id": cid})


@app.route("/api/bebidas/consumos/<cid>", methods=["DELETE"])
def delete_consumo(cid):
    if not require_permission("manage_bebidas"):
        return err("No tienes permiso para gestionar bebidas.")
    db = get_db()
    db.execute("DELETE FROM bebidas_consumos WHERE id = ?", (cid,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/bebidas/consumos/<cid>/pagado", methods=["POST"])
def toggle_consumo_pagado(cid):
    if not require_permission("manage_bebidas"):
        return err("No tienes permiso para gestionar bebidas.")
    db = get_db()
    row = db.execute("SELECT pagado FROM bebidas_consumos WHERE id = ?", (cid,)).fetchone()
    if not row:
        return err("Consumo no encontrado", 404)
    nuevo = 0 if row["pagado"] else 1
    db.execute("UPDATE bebidas_consumos SET pagado = ? WHERE id = ?", (nuevo, cid))
    db.commit()
    return jsonify({"ok": True, "pagado": bool(nuevo)})


# -------------------------------------------------------------- fiestas (gasto bebida evento) --
@app.route("/api/fiestas", methods=["POST"])
def add_fiesta_gasto():
    if not require_login():
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    db = get_db()
    fid = new_id()
    db.execute(
        "INSERT INTO fiestas_gastos (id, fecha, evento, concepto, importe) VALUES (?,?,?,?,?)",
        (
            fid,
            data.get("fecha"),
            (data.get("evento") or "").strip(),
            (data.get("concepto") or "").strip(),
            float(data.get("importe") or 0),
        ),
    )
    db.commit()
    return jsonify({"ok": True, "id": fid})


@app.route("/api/fiestas/<fid>", methods=["DELETE"])
def delete_fiesta_gasto(fid):
    if not require_login():
        return err("No has iniciado sesion.", 401)
    db = get_db()
    db.execute("DELETE FROM fiestas_gastos WHERE id = ?", (fid,))
    db.commit()
    return jsonify({"ok": True})


# -------------------------------------------------------------- reparto de gastos por evento --
@app.route("/api/gastos-eventos", methods=["POST"])
def add_gasto_evento():
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return err("El nombre del evento es obligatorio.", 400)
    db = get_db()
    validos = {r["id"] for r in db.execute("SELECT id FROM socios")}
    participantes = data.get("participantes") or []
    if not isinstance(participantes, list):
        participantes = []
    # Quien crea el evento no participa automaticamente: se anade a la lista
    # de participantes solo si el se ha marcado a si mismo, igual que el resto.
    participantes = [p for p in participantes if p in validos]
    eid = new_id()
    db.execute(
        "INSERT INTO gastos_eventos (id, nombre, fecha, notas, creado_por, creado_en) VALUES (?,?,?,?,?,?)",
        (eid, nombre, data.get("fecha") or date.today().isoformat(), (data.get("notas") or "").strip(), sid, now_iso()),
    )
    db.executemany(
        "INSERT OR IGNORE INTO gastos_evento_participantes (evento_id, socio_id) VALUES (?, ?)",
        [(eid, p) for p in participantes],
    )
    db.commit()
    return jsonify({"ok": True, "id": eid})


@app.route("/api/gastos-eventos/<eid>", methods=["DELETE"])
def delete_gasto_evento(eid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    row = db.execute("SELECT creado_por FROM gastos_eventos WHERE id = ?", (eid,)).fetchone()
    if not row:
        return err("Evento no encontrado", 404)
    if row["creado_por"] != sid and not has_permission(sid, "manage_finances"):
        return err("Solo quien creo el evento (o quien gestiona finanzas) puede borrarlo.")
    db.execute("DELETE FROM gastos_evento_pagos WHERE evento_id = ?", (eid,))
    db.execute("DELETE FROM gastos_evento_participantes WHERE evento_id = ?", (eid,))
    db.execute("DELETE FROM gastos_eventos WHERE id = ?", (eid,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/gastos-eventos/<eid>/ocultar", methods=["POST"])
def toggle_gasto_evento_oculto(eid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    row = db.execute("SELECT creado_por, oculto FROM gastos_eventos WHERE id = ?", (eid,)).fetchone()
    if not row:
        return err("Evento no encontrado", 404)
    if row["creado_por"] != sid and not has_permission(sid, "manage_finances"):
        return err("Solo quien creo el evento (o quien gestiona finanzas) puede ocultarlo.")
    nuevo = 0 if row["oculto"] else 1
    db.execute("UPDATE gastos_eventos SET oculto = ? WHERE id = ?", (nuevo, eid))
    db.commit()
    return jsonify({"ok": True, "oculto": bool(nuevo)})


@app.route("/api/gastos-eventos/<eid>/participantes/toggle", methods=["POST"])
def toggle_gasto_evento_participante(eid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    target = data.get("socio_id")
    db = get_db()
    if not db.execute("SELECT 1 FROM gastos_eventos WHERE id = ?", (eid,)).fetchone():
        return err("Evento no encontrado", 404)
    if not db.execute("SELECT 1 FROM socios WHERE id = ?", (target,)).fetchone():
        return err("Socio no encontrado", 404)
    ya = db.execute(
        "SELECT 1 FROM gastos_evento_participantes WHERE evento_id = ? AND socio_id = ?", (eid, target)
    ).fetchone()
    if ya:
        tiene_pagos = db.execute(
            "SELECT 1 FROM gastos_evento_pagos WHERE evento_id = ? AND pagador_id = ?", (eid, target)
        ).fetchone()
        if tiene_pagos:
            return err("Este socio ya tiene pagos registrados en el evento: borralos primero.", 400)
        db.execute("DELETE FROM gastos_evento_participantes WHERE evento_id = ? AND socio_id = ?", (eid, target))
    else:
        db.execute("INSERT INTO gastos_evento_participantes (evento_id, socio_id) VALUES (?, ?)", (eid, target))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/gastos-eventos/<eid>/pagos", methods=["POST"])
def add_gasto_evento_pago(eid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    if not db.execute("SELECT 1 FROM gastos_eventos WHERE id = ?", (eid,)).fetchone():
        return err("Evento no encontrado", 404)
    data = request.get_json(force=True)
    participantes = {
        r["socio_id"] for r in db.execute("SELECT socio_id FROM gastos_evento_participantes WHERE evento_id = ?", (eid,))
    }
    pagador_id = data.get("pagador_id")
    if pagador_id not in participantes:
        return err("El pagador debe ser uno de los participantes del evento.", 400)
    beneficiarios = data.get("beneficiarios") or []
    if not isinstance(beneficiarios, list):
        beneficiarios = []
    beneficiarios = [b for b in beneficiarios if b in participantes]
    try:
        importe = float(data.get("importe") or 0)
    except (TypeError, ValueError):
        importe = 0
    if importe <= 0:
        return err("El importe debe ser mayor que 0.", 400)
    concepto = (data.get("concepto") or "").strip()
    if not concepto:
        return err("El concepto es obligatorio.", 400)
    pid = new_id()
    db.execute(
        "INSERT INTO gastos_evento_pagos (id, evento_id, pagador_id, concepto, importe, fecha, beneficiarios) VALUES (?,?,?,?,?,?,?)",
        (pid, eid, pagador_id, concepto, importe, data.get("fecha") or date.today().isoformat(), json.dumps(beneficiarios)),
    )
    db.commit()
    return jsonify({"ok": True, "id": pid})


@app.route("/api/gastos-eventos/<eid>/pagos/<pid>", methods=["DELETE"])
def delete_gasto_evento_pago(eid, pid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    pago = db.execute(
        "SELECT pagador_id FROM gastos_evento_pagos WHERE id = ? AND evento_id = ?", (pid, eid)
    ).fetchone()
    if not pago:
        return err("Pago no encontrado", 404)
    evento = db.execute("SELECT creado_por FROM gastos_eventos WHERE id = ?", (eid,)).fetchone()
    puede = sid == pago["pagador_id"] or (evento and sid == evento["creado_por"]) or has_permission(sid, "manage_finances")
    if not puede:
        return err("Solo quien pago, quien creo el evento o quien gestiona finanzas puede borrar este pago.")
    db.execute("DELETE FROM gastos_evento_pagos WHERE id = ?", (pid,))
    db.commit()
    return jsonify({"ok": True})


# -------------------------------------------------------------- gastos de socios --
def _gasto_socio_o_error(sid, gid):
    db = get_db()
    row = db.execute("SELECT * FROM gastos_socios WHERE id = ?", (gid,)).fetchone()
    if not row:
        return None, err("Gasto no encontrado", 404)
    if row["socio_id"] != sid and not has_permission(sid, "manage_finances"):
        return None, err("Solo quien registro este gasto o quien gestiona finanzas puede modificarlo.")
    return row, None


@app.route("/api/gastos-socios", methods=["POST"])
def add_gasto_socio():
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    db = get_db()
    socio_id = sid
    concepto = (data.get("concepto") or "").strip()
    if not concepto:
        return err("El concepto es obligatorio.", 400)
    try:
        importe = float(data.get("importe") or 0)
    except (TypeError, ValueError):
        importe = 0
    if importe <= 0:
        return err("El importe debe ser mayor que 0.", 400)
    tipo = data.get("tipo") if data.get("tipo") in ("gasto", "ingreso") else "gasto"
    gid = new_id()
    db.execute(
        "INSERT INTO gastos_socios (id, socio_id, tipo, concepto, importe, fecha, notas, creado_en) VALUES (?,?,?,?,?,?,?,?)",
        (
            gid,
            socio_id,
            tipo,
            concepto,
            importe,
            data.get("fecha") or date.today().isoformat(),
            (data.get("notas") or "").strip(),
            now_iso(),
        ),
    )
    db.commit()
    return jsonify({"ok": True, "id": gid})


@app.route("/api/gastos-socios/<gid>", methods=["POST"])
def update_gasto_socio(gid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    row, error = _gasto_socio_o_error(sid, gid)
    if error:
        return error
    data = request.get_json(force=True)
    concepto = (data.get("concepto") or "").strip()
    if not concepto:
        return err("El concepto es obligatorio.", 400)
    try:
        importe = float(data.get("importe") or 0)
    except (TypeError, ValueError):
        importe = 0
    if importe <= 0:
        return err("El importe debe ser mayor que 0.", 400)
    tipo = data.get("tipo") if data.get("tipo") in ("gasto", "ingreso") else row["tipo"]
    db = get_db()
    db.execute(
        "UPDATE gastos_socios SET tipo=?, concepto=?, importe=?, fecha=?, notas=? WHERE id=?",
        (tipo, concepto, importe, data.get("fecha") or date.today().isoformat(), (data.get("notas") or "").strip(), gid),
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/gastos-socios/<gid>/abonado", methods=["POST"])
def toggle_abonado_gasto_socio(gid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    if not has_permission(sid, "manage_finances"):
        return err("Solo el tesorero o el administrador pueden marcarlo, una vez comprobado.")
    db = get_db()
    row = db.execute("SELECT abonado FROM gastos_socios WHERE id = ?", (gid,)).fetchone()
    if not row:
        return err("Gasto no encontrado", 404)
    db.execute("UPDATE gastos_socios SET abonado = ? WHERE id = ?", (0 if row["abonado"] else 1, gid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/gastos-socios/<gid>", methods=["DELETE"])
def delete_gasto_socio(gid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    row, error = _gasto_socio_o_error(sid, gid)
    if error:
        return error
    db = get_db()
    db.execute("DELETE FROM gastos_socios WHERE id = ?", (gid,))
    db.commit()
    for ext in ("jpg", "pdf"):
        ticket_path = os.path.join(TICKETS_DIR, f"{gid}.{ext}")
        try:
            if os.path.exists(ticket_path):
                os.remove(ticket_path)
        except OSError:
            pass
    return jsonify({"ok": True})


@app.route("/api/gastos-socios/<gid>/ticket", methods=["POST"])
def upload_ticket_gasto_socio(gid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    row, error = _gasto_socio_o_error(sid, gid)
    if error:
        return error
    if "ticket" not in request.files or request.files["ticket"].filename == "":
        return err("No se ha recibido ningun archivo.", 400)
    ext, error = _guardar_ticket_o_factura(request.files["ticket"], TICKETS_DIR, gid)
    if error:
        return error
    db = get_db()
    db.execute("UPDATE gastos_socios SET ticket = 1, ticket_ext = ? WHERE id = ?", (ext, gid))
    db.commit()
    return jsonify({"ok": True, "ext": ext})


@app.route("/api/gastos-socios/<gid>/ticket", methods=["DELETE"])
def delete_ticket_gasto_socio(gid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    row, error = _gasto_socio_o_error(sid, gid)
    if error:
        return error
    db = get_db()
    for ext in ("jpg", "pdf"):
        ticket_path = os.path.join(TICKETS_DIR, f"{gid}.{ext}")
        try:
            if os.path.exists(ticket_path):
                os.remove(ticket_path)
        except OSError:
            pass
    db.execute("UPDATE gastos_socios SET ticket = 0, ticket_ext = '' WHERE id = ?", (gid,))
    db.commit()
    return jsonify({"ok": True})


def _guardar_ticket(file, dest_path):
    try:
        raw = file.read()
        try:
            im = Image.open(BytesIO(raw))
            im.load()
        except Exception:
            return False, err(
                "No se pudo leer esa imagen. Prueba con un archivo .jpg o .png "
                "(si es una foto de iPhone en formato HEIC, conviertela a JPG antes de subirla).",
                400,
            )
        im = ImageOps.exif_transpose(im).convert("RGB")
        im.thumbnail((1600, 1600), Image.LANCZOS)
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        im.save(dest_path, "JPEG", quality=85)
    except Exception as e:
        return False, err(f"No se pudo procesar la imagen: {e}", 400)
    return True, None


def _guardar_ticket_o_factura(file, dest_dir, base_name):
    """Como _guardar_ticket, pero admite tambien PDF (facturas), sin convertirlo.
    Devuelve (extension_guardada, None) o (None, respuesta_de_error)."""
    filename = (file.filename or "").lower()
    content_type = (file.mimetype or "").lower()
    es_pdf = filename.endswith(".pdf") or content_type == "application/pdf"
    raw = file.read()
    os.makedirs(dest_dir, exist_ok=True)
    for ext_previa in ("jpg", "pdf"):
        try:
            os.remove(os.path.join(dest_dir, f"{base_name}.{ext_previa}"))
        except OSError:
            pass
    if es_pdf:
        if raw[:4] != b"%PDF":
            return None, err("El archivo no parece ser un PDF valido.", 400)
        if len(raw) > 15 * 1024 * 1024:
            return None, err("El PDF es demasiado grande (maximo 15 MB).", 400)
        with open(os.path.join(dest_dir, f"{base_name}.pdf"), "wb") as f:
            f.write(raw)
        return "pdf", None
    if not PIL_OK:
        return None, err(
            "Falta instalar una dependencia en el servidor: ejecuta "
            "'pip install -r requirements.txt' (o 'pip install Pillow') "
            "y reinicia la aplicacion.",
            500,
        )
    try:
        im = Image.open(BytesIO(raw))
        im.load()
    except Exception:
        return None, err(
            "No se pudo leer ese archivo. Sube una imagen (.jpg, .png) o un PDF "
            "(si es una foto de iPhone en formato HEIC, conviertela antes a JPG).",
            400,
        )
    im = ImageOps.exif_transpose(im).convert("RGB")
    im.thumbnail((1600, 1600), Image.LANCZOS)
    im.save(os.path.join(dest_dir, f"{base_name}.jpg"), "JPEG", quality=85)
    return "jpg", None


@app.route("/api/gastos-eventos/<eid>/pagos/<pid>/ticket", methods=["POST"])
def upload_ticket_gasto_evento_pago(eid, pid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    pago = db.execute(
        "SELECT pagador_id FROM gastos_evento_pagos WHERE id = ? AND evento_id = ?", (pid, eid)
    ).fetchone()
    if not pago:
        return err("Pago no encontrado", 404)
    evento = db.execute("SELECT creado_por FROM gastos_eventos WHERE id = ?", (eid,)).fetchone()
    puede = sid == pago["pagador_id"] or (evento and sid == evento["creado_por"]) or has_permission(sid, "manage_finances")
    if not puede:
        return err("Solo quien pago, quien creo el evento o quien gestiona finanzas puede subir el ticket de este pago.")
    if not PIL_OK:
        return err(
            "Falta instalar una dependencia en el servidor: ejecuta "
            "'pip install -r requirements.txt' (o 'pip install Pillow') "
            "y reinicia la aplicacion.",
            500,
        )
    if "ticket" not in request.files or request.files["ticket"].filename == "":
        return err("No se ha recibido ninguna imagen.", 400)
    guardado, error = _guardar_ticket(request.files["ticket"], os.path.join(TICKETS_DIR, f"pago-{pid}.jpg"))
    if error:
        return error
    db.execute("UPDATE gastos_evento_pagos SET ticket = 1 WHERE id = ?", (pid,))
    db.commit()
    return jsonify({"ok": True})


# -------------------------------------------------------------- reservas --
@app.route("/api/reservas", methods=["POST"])
def add_reserva():
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    db = get_db()
    rid = new_id()
    db.execute(
        "INSERT INTO reservas (id, fecha, hora_inicio, hora_fin, socio_id, evento, notas, creado_en) VALUES (?,?,?,?,?,?,?,?)",
        (
            rid,
            data.get("fecha"),
            (data.get("hora_inicio") or "").strip(),
            (data.get("hora_fin") or "").strip(),
            sid,
            (data.get("evento") or "").strip(),
            (data.get("notas") or "").strip(),
            now_iso(),
        ),
    )
    db.commit()
    return jsonify({"ok": True, "id": rid})


@app.route("/api/reservas/<rid>", methods=["DELETE"])
def delete_reserva(rid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    row = db.execute("SELECT socio_id FROM reservas WHERE id = ?", (rid,)).fetchone()
    if not row:
        return err("Reserva no encontrada", 404)
    if row["socio_id"] != sid and not has_permission(sid, "manage_events"):
        return err("No puedes cancelar la reserva de otro socio.")
    db.execute("DELETE FROM reservas WHERE id = ?", (rid,))
    db.commit()
    return jsonify({"ok": True})


# -------------------------------------------------------------- listas de la compra --
@app.route("/api/listas-compra", methods=["POST"])
def add_lista_compra():
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return err("El nombre de la lista es obligatorio.", 400)
    db = get_db()
    lid = new_id()
    db.execute(
        "INSERT INTO listas_compra (id, nombre, creado_por, creado_en) VALUES (?,?,?,?)",
        (lid, nombre, sid, now_iso()),
    )
    db.commit()
    return jsonify({"ok": True, "id": lid})


@app.route("/api/listas-compra/<lid>", methods=["DELETE"])
def delete_lista_compra(lid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    if not db.execute("SELECT 1 FROM listas_compra WHERE id = ?", (lid,)).fetchone():
        return err("Lista no encontrada", 404)
    db.execute("DELETE FROM lista_compra_items WHERE lista_id = ?", (lid,))
    db.execute("DELETE FROM listas_compra WHERE id = ?", (lid,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/listas-compra/<lid>/items", methods=["POST"])
def add_item_compra(lid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    if not db.execute("SELECT 1 FROM listas_compra WHERE id = ?", (lid,)).fetchone():
        return err("Lista no encontrada", 404)
    data = request.get_json(force=True)
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return err("El nombre del producto es obligatorio.", 400)
    iid = new_id()
    db.execute(
        "INSERT INTO lista_compra_items (id, lista_id, nombre, cantidad, comprado, anadido_por, creado_en) "
        "VALUES (?,?,?,?,0,?,?)",
        (iid, lid, nombre, (data.get("cantidad") or "").strip(), sid, now_iso()),
    )
    db.commit()
    return jsonify({"ok": True, "id": iid})


@app.route("/api/listas-compra/items/<iid>", methods=["POST"])
def update_item_compra(iid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    if not db.execute("SELECT 1 FROM lista_compra_items WHERE id = ?", (iid,)).fetchone():
        return err("Producto no encontrado", 404)
    data = request.get_json(force=True)
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return err("El nombre del producto es obligatorio.", 400)
    db.execute(
        "UPDATE lista_compra_items SET nombre = ?, cantidad = ? WHERE id = ?",
        (nombre, (data.get("cantidad") or "").strip(), iid),
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/listas-compra/items/<iid>/toggle", methods=["POST"])
def toggle_item_compra(iid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    row = db.execute("SELECT comprado FROM lista_compra_items WHERE id = ?", (iid,)).fetchone()
    if not row:
        return err("Producto no encontrado", 404)
    nuevo = 0 if row["comprado"] else 1
    db.execute(
        "UPDATE lista_compra_items SET comprado = ?, comprado_por = ? WHERE id = ?",
        (nuevo, sid if nuevo else None, iid),
    )
    db.commit()
    return jsonify({"ok": True, "comprado": bool(nuevo)})


@app.route("/api/listas-compra/items/<iid>", methods=["DELETE"])
def delete_item_compra(iid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    db.execute("DELETE FROM lista_compra_items WHERE id = ?", (iid,))
    db.commit()
    return jsonify({"ok": True})


# -------------------------------------------------------------- tareas: tickets --
@app.route("/api/tareas-tickets", methods=["POST"])
def add_tarea_ticket():
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    data = request.get_json(force=True)
    tipo = (data.get("tipo") or "").strip()
    if not tipo:
        return err("El tipo de tarea es obligatorio.", 400)
    socios_in = data.get("socios")
    if not isinstance(socios_in, list):
        socios_in = [data.get("responsable_id")] if data.get("responsable_id") else []
    estado = data.get("estado") or "pendiente"
    if estado not in ESTADOS_TICKET:
        estado = "pendiente"
    db = get_db()
    socios_validos = {r["id"] for r in db.execute("SELECT id FROM socios")}
    socios = []
    for s in socios_in:
        if s and s in socios_validos and s not in socios:
            socios.append(s)
    tid = new_id()
    db.execute(
        "INSERT INTO tareas_tickets (id, tipo, responsable_id, estado, fecha, notas, rotacion, creado_por, creado_en) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        (
            tid, tipo, socios[0] if socios else None, estado,
            data.get("fecha") or None,
            (data.get("notas") or "").strip(), (data.get("rotacion") or "").strip(),
            sid, now_iso(),
        ),
    )
    if socios:
        db.executemany("INSERT OR IGNORE INTO tareas_ticket_socios (ticket_id, socio_id) VALUES (?, ?)", [(tid, s) for s in socios])
    db.commit()
    return jsonify({"ok": True, "id": tid})


@app.route("/api/tareas-tickets/<tid>", methods=["POST"])
def update_tarea_ticket(tid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    ticket = db.execute("SELECT * FROM tareas_tickets WHERE id = ?", (tid,)).fetchone()
    if not ticket:
        return err("Ticket no encontrado", 404)

    data = request.get_json(force=True)
    campos, valores = [], []

    if "socios" in data:
        socios_in = data.get("socios")
        if not isinstance(socios_in, list):
            socios_in = []
        socios_validos = {r["id"] for r in db.execute("SELECT id FROM socios")}
        socios = []
        for s in socios_in:
            if s and s in socios_validos and s not in socios:
                socios.append(s)
        db.execute("DELETE FROM tareas_ticket_socios WHERE ticket_id = ?", (tid,))
        if socios:
            db.executemany("INSERT OR IGNORE INTO tareas_ticket_socios (ticket_id, socio_id) VALUES (?, ?)", [(tid, s) for s in socios])
        campos.append("responsable_id = ?")
        valores.append(socios[0] if socios else None)
    nueva_tarea_generada = False
    if "estado" in data:
        estado = data.get("estado")
        if estado not in ESTADOS_TICKET:
            return err("Estado no valido.", 400)
        campos.append("estado = ?")
        valores.append(estado)
        if estado == "hecho" and ticket["estado"] != "hecho":
            campos.append("completado_en = ?")
            valores.append(now_iso())
            if (ticket["rotacion"] or "").strip():
                crear_siguiente_ocurrencia_tarea(db, ticket)
                nueva_tarea_generada = True
        elif estado != "hecho" and ticket["estado"] == "hecho":
            campos.append("completado_en = ?")
            valores.append(None)
    if "tipo" in data:
        tipo = (data.get("tipo") or "").strip()
        if not tipo:
            return err("El tipo de tarea es obligatorio.", 400)
        campos.append("tipo = ?")
        valores.append(tipo)
    if "fecha" in data:
        campos.append("fecha = ?")
        valores.append(data.get("fecha") or None)
    if "turno" in data:
        campos.append("turno = ?")
        valores.append((data.get("turno") or "").strip())
    if "notas" in data:
        campos.append("notas = ?")
        valores.append((data.get("notas") or "").strip())
    if "rotacion" in data:
        campos.append("rotacion = ?")
        valores.append((data.get("rotacion") or "").strip())

    if campos:
        valores.append(tid)
        db.execute(f"UPDATE tareas_tickets SET {', '.join(campos)} WHERE id = ?", valores)
        db.commit()
    return jsonify({"ok": True, "nueva_tarea_generada": nueva_tarea_generada})


@app.route("/api/tareas-tickets/<tid>", methods=["DELETE"])
def delete_tarea_ticket(tid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    ticket = db.execute("SELECT * FROM tareas_tickets WHERE id = ?", (tid,)).fetchone()
    if not ticket:
        return err("Ticket no encontrado", 404)
    db.execute("DELETE FROM tareas_tickets WHERE id = ?", (tid,))
    db.execute("DELETE FROM tareas_ticket_socios WHERE ticket_id = ?", (tid,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/tareas-tickets/<tid>/socios/toggle", methods=["POST"])
def toggle_socio_tarea(tid):
    sid = require_login()
    if not sid:
        return err("No has iniciado sesion.", 401)
    db = get_db()
    if not db.execute("SELECT 1 FROM tareas_tickets WHERE id = ?", (tid,)).fetchone():
        return err("Ticket no encontrado", 404)
    data = request.get_json(force=True)
    target = data.get("socio_id")
    if not target or not db.execute("SELECT 1 FROM socios WHERE id = ?", (target,)).fetchone():
        return err("Socio no encontrado", 404)
    ya_asignado = db.execute(
        "SELECT 1 FROM tareas_ticket_socios WHERE ticket_id = ? AND socio_id = ?", (tid, target)
    ).fetchone()
    if ya_asignado:
        db.execute("DELETE FROM tareas_ticket_socios WHERE ticket_id = ? AND socio_id = ?", (tid, target))
    else:
        db.execute("INSERT INTO tareas_ticket_socios (ticket_id, socio_id) VALUES (?, ?)", (tid, target))
    restantes = [r["socio_id"] for r in db.execute("SELECT socio_id FROM tareas_ticket_socios WHERE ticket_id = ?", (tid,))]
    db.execute("UPDATE tareas_tickets SET responsable_id = ? WHERE id = ?", (restantes[0] if restantes else None, tid))
    db.commit()
    return jsonify({"ok": True, "asignado": not ya_asignado})


# -------------------------------------------------------------- exportar a Excel --
@app.route("/api/export.xlsx")
def export_excel():
    if not require_permission("export_data"):
        return err("No has iniciado sesion.", 401)
    if not OPENPYXL_OK:
        return err(
            "Falta instalar una dependencia en el servidor: ejecuta "
            "'pip install -r requirements.txt' (o 'pip install openpyxl') "
            "y reinicia la aplicacion.",
            500,
        )

    try:
        return _build_excel()
    except Exception as e:
        app.logger.exception("Error generando el Excel")
        return err(f"No se pudo generar el Excel: {e}", 500)


def _build_excel():
    db = get_db()
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="21332B", end_color="21332B", fill_type="solid")
    header_font = Font(color="F1EDE4", bold=True)

    def write_sheet(ws, headers, rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for r in rows:
            ws.append(r)
        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(h) + 4)
        ws.freeze_panes = "A2"

    socios_rows = db.execute("SELECT * FROM socios ORDER BY nombre").fetchall()
    nombre_de = {s["id"]: s["nombre"] for s in socios_rows}

    def nombres(ids):
        return ", ".join(nombre_de.get(i, i) for i in ids if i)

    ws0 = wb.active
    roles_rows = db.execute("SELECT * FROM roles ORDER BY nombre").fetchall()
    nombre_rol = {r["id"]: r["nombre"] for r in roles_rows}
    roles_de_socio = {}
    for sr in db.execute("SELECT * FROM socio_roles"):
        roles_de_socio.setdefault(sr["socio_id"], []).append(nombre_rol.get(sr["rol_id"], sr["rol_id"]))

    ws0.title = "Socios"
    write_sheet(
        ws0,
        ["ID", "Nombre", "Activo", "Roles (separados por coma)"],
        [[s["id"], s["nombre"], "Si" if s["activo"] else "No", ", ".join(roles_de_socio.get(s["id"], []))] for s in socios_rows],
    )

    ws0b = wb.create_sheet("Roles")
    write_sheet(
        ws0b,
        ["ID", "Nombre", "Permisos (separados por coma)", "Es del sistema"],
        [[r["id"], r["nombre"], ",".join(json.loads(r["permisos"] or "[]")), "Si" if r["es_sistema"] else "No"] for r in roles_rows],
    )

    ws1 = wb.create_sheet("Movimientos")
    movs = db.execute("SELECT * FROM movimientos ORDER BY fecha").fetchall()
    write_sheet(
        ws1,
        ["ID", "Fecha", "Tipo", "Categoria", "Socio ID", "Socio", "Concepto", "Importe (EUR)"],
        [[m["id"], m["fecha"], m["tipo"], m["categoria"], m["socio_id"] or "", nombre_de.get(m["socio_id"], ""), m["concepto"], m["importe"]] for m in movs],
    )

    ws2 = wb.create_sheet("Cuotas")
    cuotas = db.execute("SELECT * FROM cuotas ORDER BY year, month").fetchall()
    write_sheet(
        ws2,
        ["ID", "Socio ID", "Socio", "Ano", "Mes", "Importe (EUR)", "Pagado", "Fecha de pago"],
        [[c["id"], c["socio_id"], nombre_de.get(c["socio_id"], ""), c["year"], c["month"], c["importe"], "Si" if c["pagado"] else "No", c["fecha"] or ""] for c in cuotas],
    )

    ws3p = wb.create_sheet("Bebidas-Precios")
    precios = db.execute("SELECT * FROM bebidas_precios ORDER BY nombre").fetchall()
    nombre_bebida = {p["id"]: p["nombre"] for p in precios}
    write_sheet(
        ws3p,
        ["ID", "Nombre", "Unidad", "Precio socio (EUR)", "Precio no socio (EUR)"],
        [[p["id"], p["nombre"], p["unidad"], p["precio_socio"], p["precio_no_socio"]] for p in precios],
    )

    ws3 = wb.create_sheet("Bebidas-Consumos")
    consumos = db.execute("SELECT * FROM bebidas_consumos ORDER BY fecha").fetchall()
    write_sheet(
        ws3,
        ["ID", "Fecha", "Bebida ID", "Bebida", "Socio ID", "Consumidor", "Es socio", "Cantidad", "Importe (EUR)", "Pagado"],
        [[
            c["id"], c["fecha"], c["bebida_id"], nombre_bebida.get(c["bebida_id"], ""), c["socio_id"] or "",
            c["consumidor"], "Si" if c["es_socio"] else "No", c["cantidad"], c["importe"], "Si" if c["pagado"] else "No",
        ] for c in consumos],
    )

    ws4 = wb.create_sheet("Gastos-Fiestas")
    fiestas = db.execute("SELECT * FROM fiestas_gastos ORDER BY fecha").fetchall()
    write_sheet(
        ws4,
        ["ID", "Fecha", "Evento", "Concepto", "Importe (EUR)"],
        [[f["id"], f["fecha"], f["evento"], f["concepto"], f["importe"]] for f in fiestas],
    )

    ws4b = wb.create_sheet("Inventario")
    inventario = db.execute("SELECT * FROM inventario ORDER BY categoria, nombre").fetchall()
    write_sheet(
        ws4b,
        ["ID", "Categoria", "Nombre", "Cantidad", "Estado", "Notas"],
        [[i["id"], i["categoria"], i["nombre"], i["cantidad"], i["estado"], i["notas"] or ""] for i in inventario],
    )

    ws6 = wb.create_sheet("Reservas")
    reservas = db.execute("SELECT * FROM reservas ORDER BY fecha").fetchall()
    write_sheet(
        ws6,
        ["ID", "Fecha", "Hora inicio", "Hora fin", "Socio ID", "Socio", "Evento", "Notas"],
        [[r["id"], r["fecha"], r["hora_inicio"] or "", r["hora_fin"] or "", r["socio_id"] or "", nombre_de.get(r["socio_id"], ""), r["evento"], r["notas"] or ""] for r in reservas],
    )

    ws7 = wb.create_sheet("Reuniones")
    reuniones = db.execute("SELECT * FROM reuniones ORDER BY fecha").fetchall()
    asistentes_de = {}
    for a in db.execute("SELECT * FROM asistencia"):
        asistentes_de.setdefault(a["reunion_id"], []).append(a["socio_id"])
    write_sheet(
        ws7,
        ["ID", "Fecha", "Hora inicio", "Hora fin", "Evento", "Notas", "Asistentes"],
        [[
            r["id"], r["fecha"], r["hora_inicio"] or "", r["hora_fin"] or "", r["evento"], r["notas"] or "",
            nombres(asistentes_de.get(r["id"], [])),
        ] for r in reuniones],
    )

    ws8 = wb.create_sheet("Tareas")
    tareas = db.execute("SELECT * FROM tareas_tickets ORDER BY (fecha IS NULL), fecha, tipo").fetchall()
    tareas_socios_de = {}
    for r in db.execute("SELECT * FROM tareas_ticket_socios"):
        tareas_socios_de.setdefault(r["ticket_id"], []).append(r["socio_id"])
    write_sheet(
        ws8,
        ["ID", "Tipo", "Socios", "Estado", "Fecha", "Rotacion", "Notas", "Completado en"],
        [[
            t["id"], t["tipo"], nombres(tareas_socios_de.get(t["id"], [])), t["estado"],
            t["fecha"] or "", t["rotacion"] or "", t["notas"] or "", t["completado_en"] or "",
        ] for t in tareas],
    )

    ws9 = wb.create_sheet("Tricount-Eventos")
    eventos = db.execute("SELECT * FROM gastos_eventos ORDER BY fecha").fetchall()
    participantes_de = {}
    for p in db.execute("SELECT * FROM gastos_evento_participantes"):
        participantes_de.setdefault(p["evento_id"], []).append(p["socio_id"])
    write_sheet(
        ws9,
        ["ID", "Nombre", "Fecha", "Notas", "Participantes", "Creado por", "Oculto"],
        [[
            e["id"], e["nombre"], e["fecha"], e["notas"] or "",
            nombres(participantes_de.get(e["id"], [])), nombre_de.get(e["creado_por"], ""),
            "Si" if e["oculto"] else "No",
        ] for e in eventos],
    )

    ws10 = wb.create_sheet("Tricount-Pagos")
    pagos = db.execute("SELECT * FROM gastos_evento_pagos ORDER BY fecha").fetchall()
    nombre_evento = {e["id"]: e["nombre"] for e in eventos}
    write_sheet(
        ws10,
        ["ID", "Evento ID", "Evento", "Pagador ID", "Pagador", "Concepto", "Importe (EUR)", "Fecha", "Beneficiarios"],
        [[
            p["id"], p["evento_id"], nombre_evento.get(p["evento_id"], ""), p["pagador_id"], nombre_de.get(p["pagador_id"], ""),
            p["concepto"], p["importe"], p["fecha"], nombres(json.loads(p["beneficiarios"] or "[]")),
        ] for p in pagos],
    )

    ingresos_cuotas = sum(c["importe"] for c in cuotas if c["pagado"])
    ingresos_mov = sum(m["importe"] for m in movs if m["tipo"] == "ingreso")
    gastos_mov = sum(m["importe"] for m in movs if m["tipo"] == "gasto")
    ingresos_bebidas = sum(c["importe"] for c in consumos if c["pagado"])
    pendiente_bebidas = sum(c["importe"] for c in consumos if not c["pagado"])
    gastos_fiestas = sum(f["importe"] for f in fiestas)
    saldo = ingresos_cuotas + ingresos_mov + ingresos_bebidas - gastos_mov - gastos_fiestas

    ingresos_por_categoria = {}
    gastos_por_categoria = {}
    for m in movs:
        destino = ingresos_por_categoria if m["tipo"] == "ingreso" else gastos_por_categoria
        destino[m["categoria"]] = destino.get(m["categoria"], 0) + m["importe"]

    ws5 = wb.create_sheet("Resumen")
    ws5.append(["Concepto", "Importe (EUR)"])
    for cell in ws5[1]:
        cell.fill = header_fill
        cell.font = header_font
    filas = [("Cuotas cobradas", ingresos_cuotas)]
    filas.append(("Otros ingresos (total)", ingresos_mov))
    for categoria, importe in sorted(ingresos_por_categoria.items()):
        filas.append((f"  - {categoria}", importe))
    filas.append(("Bebidas cobradas", ingresos_bebidas))
    filas.append(("Bebidas pendientes de cobro", pendiente_bebidas))
    filas.append(("Gastos generales (total)", -gastos_mov))
    for categoria, importe in sorted(gastos_por_categoria.items()):
        filas.append((f"  - {categoria}", -importe))
    filas.append(("Gastos de fiestas", -gastos_fiestas))
    filas.append(("SALDO TOTAL (solo lo cobrado)", saldo))
    for concepto, importe in filas:
        ws5.append([concepto, importe])
    ws5.column_dimensions["A"].width = 32
    ws5.column_dimensions["B"].width = 16
    wb.move_sheet("Resumen", offset=-(len(wb.sheetnames) - 1))  # dejarla la primera

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"el_cado_datos_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# -------------------------------------------------------------- importar desde Excel --
def _cell_str(v):
    return "" if v is None else str(v).strip()


def _cell_bool(v):
    return _cell_str(v).lower() in ("si", "s", "yes", "y", "true", "1", "x")


def _cell_float(v, default=0.0):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (TypeError, ValueError):
        return default


def _cell_int(v, default=None):
    try:
        if v is None or v == "":
            return default
        return int(v)
    except (TypeError, ValueError):
        return default


def _sheet_rows(wb, name):
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]


@app.route("/api/import.xlsx", methods=["POST"])
def import_excel():
    sid = require_login()
    if not sid or not has_permission(sid, "manage_roles"):
        return err("Solo quien gestiona roles puede importar datos (afecta a socios, roles y finanzas).", 403)
    if not OPENPYXL_OK:
        return err("Falta instalar openpyxl en el servidor.", 500)
    if "archivo" not in request.files or request.files["archivo"].filename == "":
        return err("No se ha recibido ningun archivo.", 400)

    try:
        wb = openpyxl.load_workbook(BytesIO(request.files["archivo"].read()), data_only=True)
    except Exception as e:
        return err(f"No se pudo leer el archivo. Asegurate de que es un .xlsx valido: {e}", 400)

    try:
        resumen = _import_excel(wb, sid)
    except Exception as e:
        app.logger.exception("Error importando el Excel")
        return err(f"No se pudo importar: {e}", 500)
    return jsonify({"ok": True, "resumen": resumen})


def _import_excel(wb, sid):
    db = get_db()
    resumen = {}

    def marcar(hoja, creado=False, actualizado=False, error=None):
        r = resumen.setdefault(hoja, {"creados": 0, "actualizados": 0, "errores": []})
        if creado:
            r["creados"] += 1
        if actualizado:
            r["actualizados"] += 1
        if error:
            r["errores"].append(error)

    socios_by_id = {s["id"] for s in db.execute("SELECT id FROM socios")}
    socios_by_name = {s["nombre"].strip().lower(): s["id"] for s in db.execute("SELECT id, nombre FROM socios")}

    def resolver_socio(id_val, nombre_val):
        id_val = _cell_str(id_val)
        if id_val and id_val in socios_by_id:
            return id_val
        nombre_val = _cell_str(nombre_val)
        if nombre_val and nombre_val.lower() in socios_by_name:
            return socios_by_name[nombre_val.lower()]
        return None

    def resolver_socios(texto):
        nombres = [n.strip() for n in _cell_str(texto).split(",") if n.strip()]
        return [socios_by_name[n.lower()] for n in nombres if n.lower() in socios_by_name]

    # ---- Socios ----
    pendientes_roles = []  # (socio_id, [nombres de rol])
    for row in _sheet_rows(wb, "Socios"):
        rid, nombre, activo, roles_txt = (list(row) + [None] * 4)[:4]
        nombre = _cell_str(nombre)
        if not nombre:
            marcar("Socios", error="Fila sin nombre, omitida.")
            continue
        rid = _cell_str(rid)
        target_id = rid if rid in socios_by_id else socios_by_name.get(nombre.lower())
        activo_val = 1 if (activo is None or _cell_bool(activo)) else 0
        if target_id:
            db.execute("UPDATE socios SET nombre = ?, activo = ? WHERE id = ?", (nombre, activo_val, target_id))
            marcar("Socios", actualizado=True)
        else:
            target_id = rid if rid and rid not in socios_by_id else new_id()
            db.execute(
                "INSERT INTO socios (id, nombre, is_admin, activo, pin_hash, must_change_pin) VALUES (?, ?, 0, ?, NULL, 0)",
                (target_id, nombre, activo_val),
            )
            db.execute("INSERT INTO perfiles (socio_id, telefono, notas) VALUES (?, '', '')", (target_id,))
            db.execute("INSERT OR IGNORE INTO socio_roles (socio_id, rol_id) VALUES (?, 'socio')", (target_id,))
            marcar("Socios", creado=True)
        socios_by_id.add(target_id)
        socios_by_name[nombre.lower()] = target_id
        if roles_txt is not None:
            pendientes_roles.append((target_id, [r.strip() for r in _cell_str(roles_txt).split(",") if r.strip()]))

    # ---- Roles ----
    roles_by_id = {r["id"] for r in db.execute("SELECT id FROM roles")}
    roles_by_name = {r["nombre"].strip().lower(): r["id"] for r in db.execute("SELECT id, nombre FROM roles")}
    for row in _sheet_rows(wb, "Roles"):
        rid, nombre, permisos_txt, _es_sistema = (list(row) + [None] * 4)[:4]
        nombre = _cell_str(nombre)
        if not nombre:
            marcar("Roles", error="Fila sin nombre, omitida.")
            continue
        permisos = [p.strip() for p in _cell_str(permisos_txt).split(",") if p.strip() in PERMISSIONS]
        rid = _cell_str(rid)
        target_id = rid if rid in roles_by_id else roles_by_name.get(nombre.lower())
        if target_id:
            db.execute("UPDATE roles SET permisos = ? WHERE id = ?", (json.dumps(permisos), target_id))
            marcar("Roles", actualizado=True)
        else:
            target_id = new_id()
            try:
                db.execute("INSERT INTO roles (id, nombre, permisos, es_sistema) VALUES (?, ?, ?, 0)", (target_id, nombre, json.dumps(permisos)))
                marcar("Roles", creado=True)
            except sqlite3.IntegrityError:
                marcar("Roles", error=f"Ya existe un rol llamado '{nombre}'.")
                continue
        roles_by_id.add(target_id)
        roles_by_name[nombre.lower()] = target_id

    for socio_id, nombres_rol in pendientes_roles:
        rol_ids = [roles_by_name[n.lower()] for n in nombres_rol if n.lower() in roles_by_name]
        if rol_ids:
            db.execute("DELETE FROM socio_roles WHERE socio_id = ?", (socio_id,))
            db.executemany("INSERT INTO socio_roles (socio_id, rol_id) VALUES (?, ?)", [(socio_id, r) for r in rol_ids])
            db.execute("UPDATE socios SET is_admin = ? WHERE id = ?", (int("administrador" in rol_ids), socio_id))

    # ---- Bebidas-Precios ----
    bebidas_by_id = {b["id"] for b in db.execute("SELECT id FROM bebidas_precios")}
    bebidas_by_name = {b["nombre"].strip().lower(): b["id"] for b in db.execute("SELECT id, nombre FROM bebidas_precios")}
    for row in _sheet_rows(wb, "Bebidas-Precios"):
        rid, nombre, unidad, p_socio, p_no_socio = (list(row) + [None] * 5)[:5]
        nombre = _cell_str(nombre)
        if not nombre:
            marcar("Bebidas-Precios", error="Fila sin nombre, omitida.")
            continue
        rid = _cell_str(rid)
        target_id = rid if rid in bebidas_by_id else bebidas_by_name.get(nombre.lower())
        valores = (nombre, _cell_str(unidad), _cell_float(p_socio), _cell_float(p_no_socio))
        if target_id:
            db.execute("UPDATE bebidas_precios SET nombre=?, unidad=?, precio_socio=?, precio_no_socio=? WHERE id=?", valores + (target_id,))
            marcar("Bebidas-Precios", actualizado=True)
        else:
            target_id = new_id()
            db.execute("INSERT INTO bebidas_precios (id, nombre, unidad, precio_socio, precio_no_socio) VALUES (?,?,?,?,?)", (target_id,) + valores)
            marcar("Bebidas-Precios", creado=True)
        bebidas_by_id.add(target_id)
        bebidas_by_name[nombre.lower()] = target_id

    # ---- Tricount-Eventos ----
    eventos_by_id = {e["id"] for e in db.execute("SELECT id FROM gastos_eventos")}
    eventos_by_key = {(e["nombre"].strip().lower(), e["fecha"]): e["id"] for e in db.execute("SELECT id, nombre, fecha FROM gastos_eventos")}
    eventos_by_name_import = {}
    for row in _sheet_rows(wb, "Tricount-Eventos"):
        rid, nombre, fecha, notas, participantes_txt, creado_por_txt, oculto_txt = (list(row) + [None] * 7)[:7]
        nombre = _cell_str(nombre)
        fecha = _cell_str(fecha) or date.today().isoformat()
        if not nombre:
            marcar("Tricount-Eventos", error="Fila sin nombre, omitida.")
            continue
        rid = _cell_str(rid)
        target_id = rid if rid in eventos_by_id else eventos_by_key.get((nombre.lower(), fecha))
        creado_por = resolver_socio(None, creado_por_txt) or sid
        oculto_val = 1 if _cell_bool(oculto_txt) else 0
        if target_id:
            db.execute("UPDATE gastos_eventos SET nombre=?, fecha=?, notas=?, oculto=? WHERE id=?", (nombre, fecha, _cell_str(notas), oculto_val, target_id))
            marcar("Tricount-Eventos", actualizado=True)
        else:
            target_id = new_id()
            db.execute(
                "INSERT INTO gastos_eventos (id, nombre, fecha, notas, creado_por, creado_en, oculto) VALUES (?,?,?,?,?,?,?)",
                (target_id, nombre, fecha, _cell_str(notas), creado_por, now_iso(), oculto_val),
            )
            marcar("Tricount-Eventos", creado=True)
        participantes = resolver_socios(participantes_txt)
        if creado_por and creado_por not in participantes:
            participantes.append(creado_por)
        db.execute("DELETE FROM gastos_evento_participantes WHERE evento_id = ?", (target_id,))
        db.executemany(
            "INSERT OR IGNORE INTO gastos_evento_participantes (evento_id, socio_id) VALUES (?, ?)",
            [(target_id, p) for p in participantes],
        )
        eventos_by_id.add(target_id)
        eventos_by_key[(nombre.lower(), fecha)] = target_id
        eventos_by_name_import[nombre.lower()] = target_id

    # ---- Cuotas ----
    cuotas_by_id = {c["id"] for c in db.execute("SELECT id FROM cuotas")}
    cuotas_by_key = {(c["socio_id"], c["year"], c["month"]): c["id"] for c in db.execute("SELECT id, socio_id, year, month FROM cuotas")}
    for row in _sheet_rows(wb, "Cuotas"):
        rid, socio_id_txt, socio_txt, anio, mes, importe, pagado, fecha_pago = (list(row) + [None] * 8)[:8]
        socio_id = resolver_socio(socio_id_txt, socio_txt)
        anio, mes = _cell_int(anio), _cell_int(mes)
        if not socio_id or not anio or not mes:
            marcar("Cuotas", error=f"No se pudo resolver socio/ano/mes en una fila (socio={_cell_str(socio_txt)}).")
            continue
        rid = _cell_str(rid)
        target_id = rid if rid in cuotas_by_id else cuotas_by_key.get((socio_id, anio, mes))
        pagado_val = 1 if _cell_bool(pagado) else 0
        valores = (socio_id, anio, mes, _cell_float(importe), pagado_val, _cell_str(fecha_pago) or None)
        if target_id:
            db.execute("UPDATE cuotas SET socio_id=?, year=?, month=?, importe=?, pagado=?, fecha=? WHERE id=?", valores + (target_id,))
            marcar("Cuotas", actualizado=True)
        else:
            target_id = new_id()
            db.execute("INSERT INTO cuotas (id, socio_id, year, month, importe, pagado, fecha) VALUES (?,?,?,?,?,?,?)", (target_id,) + valores)
            marcar("Cuotas", creado=True)
        cuotas_by_id.add(target_id)
        cuotas_by_key[(socio_id, anio, mes)] = target_id

    # ---- Movimientos ----
    movs_by_id = {m["id"] for m in db.execute("SELECT id FROM movimientos")}
    for row in _sheet_rows(wb, "Movimientos"):
        rid, fecha, tipo, categoria, socio_id_txt, socio_txt, concepto, importe = (list(row) + [None] * 8)[:8]
        rid = _cell_str(rid)
        socio_id = resolver_socio(socio_id_txt, socio_txt)
        valores = (
            _cell_str(fecha) or date.today().isoformat(), _cell_str(tipo) or "gasto", _cell_str(categoria) or "Otros",
            _cell_str(concepto), _cell_float(importe), socio_id,
        )
        if rid in movs_by_id:
            db.execute("UPDATE movimientos SET fecha=?, tipo=?, categoria=?, concepto=?, importe=?, socio_id=? WHERE id=?", valores + (rid,))
            marcar("Movimientos", actualizado=True)
        else:
            target_id = rid or new_id()
            db.execute("INSERT INTO movimientos (id, fecha, tipo, categoria, concepto, importe, socio_id) VALUES (?,?,?,?,?,?,?)",
                       (target_id, valores[0], valores[1], valores[2], valores[3], valores[4], valores[5]))
            movs_by_id.add(target_id)
            marcar("Movimientos", creado=True)

    # ---- Bebidas-Consumos ----
    consumos_by_id = {c["id"] for c in db.execute("SELECT id FROM bebidas_consumos")}
    for row in _sheet_rows(wb, "Bebidas-Consumos"):
        rid, fecha, bebida_id_txt, bebida_txt, socio_id_txt, consumidor, es_socio, cantidad, importe, pagado = (list(row) + [None] * 10)[:10]
        bebida_id = _cell_str(bebida_id_txt)
        if bebida_id not in bebidas_by_id:
            bebida_id = bebidas_by_name.get(_cell_str(bebida_txt).lower())
        es_socio_val = 1 if _cell_bool(es_socio) else 0
        socio_id = resolver_socio(socio_id_txt, consumidor) if es_socio_val else None
        rid = _cell_str(rid)
        valores = (
            _cell_str(fecha) or date.today().isoformat(), _cell_str(consumidor) or "Socio", es_socio_val, socio_id,
            bebida_id, _cell_int(cantidad, 1), _cell_float(importe), 1 if _cell_bool(pagado) else 0,
        )
        if rid in consumos_by_id:
            db.execute(
                "UPDATE bebidas_consumos SET fecha=?, consumidor=?, es_socio=?, socio_id=?, bebida_id=?, cantidad=?, importe=?, pagado=? WHERE id=?",
                valores + (rid,),
            )
            marcar("Bebidas-Consumos", actualizado=True)
        else:
            target_id = rid or new_id()
            db.execute(
                "INSERT INTO bebidas_consumos (id, fecha, consumidor, es_socio, socio_id, bebida_id, cantidad, importe, pagado) VALUES (?,?,?,?,?,?,?,?,?)",
                (target_id,) + valores,
            )
            consumos_by_id.add(target_id)
            marcar("Bebidas-Consumos", creado=True)

    # ---- Gastos-Fiestas ----
    fiestas_by_id = {f["id"] for f in db.execute("SELECT id FROM fiestas_gastos")}
    for row in _sheet_rows(wb, "Gastos-Fiestas"):
        rid, fecha, evento, concepto, importe = (list(row) + [None] * 5)[:5]
        rid = _cell_str(rid)
        valores = (_cell_str(fecha) or date.today().isoformat(), _cell_str(evento), _cell_str(concepto), _cell_float(importe))
        if rid in fiestas_by_id:
            db.execute("UPDATE fiestas_gastos SET fecha=?, evento=?, concepto=?, importe=? WHERE id=?", valores + (rid,))
            marcar("Gastos-Fiestas", actualizado=True)
        else:
            target_id = rid or new_id()
            db.execute("INSERT INTO fiestas_gastos (id, fecha, evento, concepto, importe) VALUES (?,?,?,?,?)", (target_id,) + valores)
            fiestas_by_id.add(target_id)
            marcar("Gastos-Fiestas", creado=True)

    # ---- Inventario ----
    inventario_by_id = {i["id"] for i in db.execute("SELECT id FROM inventario")}
    for row in _sheet_rows(wb, "Inventario"):
        rid, categoria, nombre, cantidad, estado, notas = (list(row) + [None] * 6)[:6]
        nombre = _cell_str(nombre)
        if not nombre:
            marcar("Inventario", error="Fila sin nombre, omitida.")
            continue
        rid = _cell_str(rid)
        valores = (_cell_str(categoria) or "Otros", nombre, _cell_int(cantidad, 1), _cell_str(estado) or "Bien", _cell_str(notas))
        if rid in inventario_by_id:
            db.execute("UPDATE inventario SET categoria=?, nombre=?, cantidad=?, estado=?, notas=? WHERE id=?", valores + (rid,))
            marcar("Inventario", actualizado=True)
        else:
            target_id = rid or new_id()
            db.execute("INSERT INTO inventario (id, categoria, nombre, cantidad, estado, notas) VALUES (?,?,?,?,?,?)", (target_id,) + valores)
            inventario_by_id.add(target_id)
            marcar("Inventario", creado=True)

    # ---- Reservas ----
    reservas_by_id = {r["id"] for r in db.execute("SELECT id FROM reservas")}
    for row in _sheet_rows(wb, "Reservas"):
        rid, fecha, hora_inicio, hora_fin, socio_id_txt, socio_txt, evento, notas = (list(row) + [None] * 8)[:8]
        socio_id = resolver_socio(socio_id_txt, socio_txt) or sid
        rid = _cell_str(rid)
        valores = (_cell_str(fecha) or date.today().isoformat(), _cell_str(hora_inicio), _cell_str(hora_fin), socio_id, _cell_str(evento), _cell_str(notas))
        if rid in reservas_by_id:
            db.execute("UPDATE reservas SET fecha=?, hora_inicio=?, hora_fin=?, socio_id=?, evento=?, notas=? WHERE id=?", valores + (rid,))
            marcar("Reservas", actualizado=True)
        else:
            target_id = rid or new_id()
            db.execute(
                "INSERT INTO reservas (id, fecha, hora_inicio, hora_fin, socio_id, evento, notas, creado_en) VALUES (?,?,?,?,?,?,?,?)",
                (target_id,) + valores + (now_iso(),),
            )
            reservas_by_id.add(target_id)
            marcar("Reservas", creado=True)

    # ---- Reuniones ----
    reuniones_by_id = {r["id"] for r in db.execute("SELECT id FROM reuniones")}
    for row in _sheet_rows(wb, "Reuniones"):
        rid, fecha, hora_inicio, hora_fin, evento, notas, asistentes_txt = (list(row) + [None] * 7)[:7]
        rid = _cell_str(rid)
        valores = (_cell_str(fecha) or date.today().isoformat(), _cell_str(hora_inicio), _cell_str(hora_fin), _cell_str(evento), _cell_str(notas))
        if rid in reuniones_by_id:
            db.execute("UPDATE reuniones SET fecha=?, hora_inicio=?, hora_fin=?, evento=?, notas=? WHERE id=?", valores + (rid,))
            target_id = rid
            marcar("Reuniones", actualizado=True)
        else:
            target_id = rid or new_id()
            db.execute(
                "INSERT INTO reuniones (id, fecha, hora_inicio, hora_fin, evento, notas, creado_en) VALUES (?,?,?,?,?,?,?)",
                (target_id,) + valores + (now_iso(),),
            )
            reuniones_by_id.add(target_id)
            marcar("Reuniones", creado=True)
        asistentes = resolver_socios(asistentes_txt)
        db.execute("DELETE FROM asistencia WHERE reunion_id = ?", (target_id,))
        db.executemany("INSERT OR IGNORE INTO asistencia (reunion_id, socio_id) VALUES (?, ?)", [(target_id, a) for a in asistentes])

    # ---- Tareas ----
    tareas_by_id = {t["id"] for t in db.execute("SELECT id FROM tareas_tickets")}
    for row in _sheet_rows(wb, "Tareas"):
        rid, tipo, socios_txt, estado, fecha, rotacion, notas, completado_en = (list(row) + [None] * 8)[:8]
        tipo = _cell_str(tipo)
        if not tipo:
            marcar("Tareas", error="Fila sin tipo, omitida.")
            continue
        estado = _cell_str(estado).lower() or "pendiente"
        if estado not in ESTADOS_TICKET:
            estado = "pendiente"
        socios = resolver_socios(socios_txt)
        rid = _cell_str(rid)
        valores = (
            tipo, socios[0] if socios else None, estado, _cell_str(fecha) or None,
            _cell_str(rotacion), _cell_str(notas), _cell_str(completado_en) or None,
        )
        if rid in tareas_by_id:
            db.execute(
                "UPDATE tareas_tickets SET tipo=?, responsable_id=?, estado=?, fecha=?, rotacion=?, notas=?, completado_en=? WHERE id=?",
                valores + (rid,),
            )
            target_id = rid
            marcar("Tareas", actualizado=True)
        else:
            target_id = rid or new_id()
            db.execute(
                "INSERT INTO tareas_tickets (id, tipo, responsable_id, estado, fecha, rotacion, notas, completado_en, creado_por, creado_en) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (target_id,) + valores + (sid, now_iso()),
            )
            tareas_by_id.add(target_id)
            marcar("Tareas", creado=True)
        db.execute("DELETE FROM tareas_ticket_socios WHERE ticket_id = ?", (target_id,))
        if socios:
            db.executemany("INSERT OR IGNORE INTO tareas_ticket_socios (ticket_id, socio_id) VALUES (?, ?)", [(target_id, s) for s in socios])

    # ---- Tricount-Pagos ----
    pagos_by_id = {p["id"] for p in db.execute("SELECT id FROM gastos_evento_pagos")}
    for row in _sheet_rows(wb, "Tricount-Pagos"):
        rid, evento_id_txt, evento_txt, pagador_id_txt, pagador_txt, concepto, importe, fecha, beneficiarios_txt = (list(row) + [None] * 9)[:9]
        evento_id = _cell_str(evento_id_txt)
        if evento_id not in eventos_by_id:
            evento_id = eventos_by_name_import.get(_cell_str(evento_txt).lower()) or eventos_by_key.get((_cell_str(evento_txt).lower(), _cell_str(fecha)))
        pagador_id = resolver_socio(pagador_id_txt, pagador_txt)
        if not evento_id or not pagador_id:
            marcar("Tricount-Pagos", error=f"No se pudo resolver evento/pagador en una fila (evento={_cell_str(evento_txt)}).")
            continue
        beneficiarios = resolver_socios(beneficiarios_txt)
        rid = _cell_str(rid)
        valores = (evento_id, pagador_id, _cell_str(concepto), _cell_float(importe), _cell_str(fecha) or date.today().isoformat(), json.dumps(beneficiarios))
        if rid in pagos_by_id:
            db.execute("UPDATE gastos_evento_pagos SET evento_id=?, pagador_id=?, concepto=?, importe=?, fecha=?, beneficiarios=? WHERE id=?", valores + (rid,))
            marcar("Tricount-Pagos", actualizado=True)
        else:
            target_id = rid or new_id()
            db.execute(
                "INSERT INTO gastos_evento_pagos (id, evento_id, pagador_id, concepto, importe, fecha, beneficiarios) VALUES (?,?,?,?,?,?,?)",
                (target_id,) + valores,
            )
            pagos_by_id.add(target_id)
            marcar("Tricount-Pagos", creado=True)

    db.commit()
    return resumen


if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="0.0.0.0", port=5000)
else:
    init_db()
